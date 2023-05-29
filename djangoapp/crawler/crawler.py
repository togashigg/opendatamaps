#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# crawler.py: オープンデータを取得して共通形式でキャッシュする
# Copyright (C) N.Togashi 2023

import os
import sys
import datetime
import time
import re
import csv
import json
import pathlib
import requests
from charset_normalizer import detect
import unicodedata
import openpyxl
import xlrd
from bs4 import BeautifulSoup
import logging

MY_CONFIG = {
    'download_dir': 'download',
    'cache_dir': 'cache',
    'dir_name': '{code}_{name}'
}
OPENDATA_SITES = [
    {'name': '東京都', 'code': '130001', 'api_url': 'https://catalog.data.metro.tokyo.lg.jp/api/3/action/'},
    {'name': '静岡県', 'code': '220001', 'api_url': 'https://opendata.pref.shizuoka.jp/api/'}
]
LOCALITY_CODE_FILE = '都道府県コード及び市区町村コード_20190501.csv'
IGNORE_WORD_IN_DATASET_NAME = ['台帳','統計','人口','カレンダー','コロナ','登録簿']
KIND_LIST_NORMALIZED = {
    None: re.compile('(クリーニング所|毒物劇物販売業|オープンデータ一覧|台帳|統計|人口|カレンダー|コロナ|登録簿)'),
    'AED設置箇所': re.compile('AED'),
    '介護サービス事業所': re.compile('介護'),
    '医療機関': re.compile('(病院|医療[^品]|医院|歯科|助産所|健診|応急救護|施術所|診療所)'),
    '文化財': re.compile('文化財'),
    '観光施設・場所': re.compile('(観光(施設|場所|情報|マップ)|名所|眺望|見所|ブランド|るるぶ)'),
    '公衆無線LANアクセスポイント': re.compile('((公衆|公共)?無線(LAN|ＬＡＮ)|公衆無線|Wi\-Fi|WiFi)'),
    '公衆トイレ': re.compile('(トイレ|便所)'),
    '消防水利施設': re.compile('(消防水利施設|消火栓|防火水槽)'),
    '指定緊急避難場所': re.compile('(津波|緊急避難)'),
    '公共施設': re.compile('((市|区|町|村)役所|(都|道|府|県|市|区|町|村)(の|内)(施設|機関)|庁舎|公共施設|施設情報|文化|教養|スポーツ|公民館|集会所|公会堂|(都|道|府|県|市|区|町|村)民会館|図書館|文化施設|(都|道|府|県|市|区|町|村)営住宅|斎場|墓地|環境施設|焼却施設|し尿処理|衛生検査)'),
    '子育て施設': re.compile('子育て'),
    '学校・保育施設': re.compile('(学校|こども園|幼稚園|保育|児童館|保育施設|保育所|放課後)'),
    '薬局': re.compile('(薬局|医薬品|医療品)'),
    '駐車場': re.compile('駐車場'),
    '公園・花壇': re.compile('(公園|花壇)'),
    '公衆浴場': re.compile('公衆浴場'),
    '防災': re.compile('(防災|救護所|同報無線|飲料水|ヨウ素剤|ため池|河川カメラ)'),
    '避難所': re.compile('避難(所|地|場所)'),
    '消防': re.compile('消防(署|団|施設)'),
    '投票所': re.compile('投票所'),
    '福祉施設': re.compile('(老人ホーム|生活支援ハウス|交流センター|高齢者相談センター)'),
    '健康': re.compile('健康'),
    '飲食店・販売店': re.compile('(認定店|飲食店|直売所)')
}
HEADER_ROWS = 3
LINKDATA_URL_BASE = 'http://linkdata.org'
LINKDATA_DOWNLOAD_PAGE = """<!DOCTYPE html>
<!--

    _/  _/            _/              _/              _/
   _/      _/_/_/    _/  _/      _/_/_/    _/_/_/  _/_/_/_/    _/_/_/        _/_/    _/  _/_/    _/_/_/
  _/  _/  _/    _/  _/_/      _/    _/  _/    _/    _/      _/    _/      _/    _/  _/_/      _/    _/
 _/  _/  _/    _/  _/  _/    _/    _/  _/    _/    _/      _/    _/      _/    _/  _/        _/    _/
_/  _/  _/    _/  _/    _/    _/_/_/    _/_/_/      _/_/    _/_/_/  _/    _/_/    _/          _/_/_/
                                                                                                 _/
                                                                                            _/_/
Link and Publish your data as RDF to the Linked Open Data Community
-->"""
XHTML_PAGE = '<!DOCTYPE html>'
RE_FLOAT_FORMAT = re.compile('([0-9]+\.[0-9]+)')

class Crawler:

    APP_DIR = None
    state = None        # 都道府県名
    name = None         # 市区町村名
    code = None         # 市区町村コード
    names = []
    download_dir = None
    cache_dir = None
    locality_dict = {}
    __data_name_dict = {}
    __requests_retry_max = 3
    __requests_retry_seconds = 5
    __request_headers = {}
    __proxies = {}
    __geocode_url = 'https://maps.googleapis.com/maps/api/geocode/json?address={address}&language=ja&components=country:JP&key={key}'
    __google_maps_api_key = 'AIzaSyBZa9fI3N-L1OHnkiaGQODmOcPRP-HaWlA'
    __geocode_cache = None
    __geocode_cache_file = '.geocode_cache_{}.json'
    __geocode_cache_path = None
    __geocode_cache_update = False

    def __init__(self):
        """
        コンストラクタ
        :param index_path: index.jsonファイルのパス
        :return: なし
        :raise: Exception
        """
        logger = logging.getLogger(__name__)
        logger.debug('__init__() start.')
        if 'BASE_DIR' not in locals():
            BASE_DIR=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        self.APP_DIR=os.path.join(BASE_DIR, 'djangoapp')
        self.site_names = [v['name'] for v in OPENDATA_SITES]
        logger.debug('site_names=' + str(self.site_names))
        # download_dirの確認
        self.download_dir = os.path.join(self.APP_DIR, MY_CONFIG['download_dir'])
        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir)
            logger.info('makedirs download_dir ' + self.download_dir)
        # cache_dirの確認
        self.cache_dir = os.path.join(self.APP_DIR, MY_CONFIG['cache_dir'])
        if not os.path.exists(self.cache_dir):
            os.makedirs(self.cache_dir)
            logger.info('makedirs cache_dir ' + self.cache_dir)
        logger.debug('__init__() ended.')
        # 復帰
        return

    def __del__(self):
        """
        デストラクタ
        :return: なし
        """
        logger = logging.getLogger(__name__)
        logger.debug('__del__() start.')
        pass
        logger.debug('__del__() ended.')
        return

    def get_names(self):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('get_names() start.')
        # 復帰
        logger.debug('get_names() ended, rc=' + str(self.site_names))
        return self.site_names

    def make_cache_data(self, name):
        """
        指定された自治体のオープンデータを取得しマッピング形式に変換してキャッシュする
        :param name: str型、自治体名
        :return: なし
        :raise: Exception
        """
        logger = logging.getLogger(__name__)
        logger.debug('make_cache_data() start, name=' + str(name))
        # 実行
        rc = 0
        try:
            # 自治体名の定義確認
            self.name = name
            if self.name not in self.site_names:
                raise Exception('name not exists in index, ' + self.name)
            self.state = [v['state'] for v in self.index['list'] if v['name']==self.name][0]
            self.code = [v['code'] for v in self.index['list'] if v['name']==self.name][0]
            logger.debug('code=' + self.code)
            # download_dirを設定
            self.download_dir = os.path.join(self.APP_DIR,
                    MY_CONFIG['download_dir'],
                    MY_CONFIG['dir_name'].format(
                            code=self.code, name=self.state+self.name))
            if not os.path.exists(self.download_dir):
                # raise Exception('download_dir not exists, ' + self.download_dir)
                os.mkdir(self.download_dir)
                logger.info('mkdir download_dir ' + self.download_dir)
            logger.debug('download_dir=' + self.download_dir)
            # cache_dirを設定
            self.cache_dir = os.path.join(self.APP_DIR,
                    MY_CONFIG['cache_dir'],
                    MY_CONFIG['dir_name'].format(
                            code=self.code, name=self.state+self.name))
            if not os.path.exists(self.cache_dir):
                # raise Exception('cache_dir not exists, ' + self.cache_dir)
                os.mkdir(self.cache_dir)
                logger.info('mkdir cache_dir ' + self.cache_dir)
            logger.debug('cache_dir=' + self.cache_dir)
            # geocode検索結果キャッシュを初期化
            self.__geocode_cache = {}
            self.__geocode_cache_update = False
            self.__geocode_cache_path = os.path.join(self.cache_dir, self.__geocode_cache_file.format(name))
            if os.path.exists(self.__geocode_cache_path):
                with (open(self.__geocode_cache_path, 'r') as hf):
                    self.__geocode_cache = json.loads(hf.read())
            # 情報一覧取得
            titles = [y['title'] for x in self.index['list'] if x['name']==self.name for y in x['data']]
            logger.debug('titles=' + str(titles))

            info_list = [v['data'] for v in self.index['list'] if v['name']==self.name][0]
            for info in info_list:
                file = info['url'].split('/')[-1]
                file_cache = str(pathlib.Path(file).with_suffix('.json'))
                if self.exists_in_cache(file_cache):
                    msg = info['dataset'] + '：' + file_cache + ' 取得済'
                    logger.info(msg)
                    print(msg, file=sys.stderr)
                else:
                    msg = info['dataset'] + '：' + file + ' 取得開始'
                    logger.info(msg)
                    print(msg, file=sys.stderr, end=' ... ')
                    content = self.url_get(info['url'], file)
                    if content is None:
                        print('データ取得失敗', file=sys.stderr)
                    else:
                        print('変換開始', file=sys.stderr, end=' ... ')
                        if info['format'] in ['CSV', 'TEXT', 'TXT', 'TSV']:
                            content = self.convert_to_utf8(content)
                        if content is None:
                            print('形式誤り', file=sys.stderr)
                        else:
                            download_path = os.path.join(self.download_dir, file)
                            map_data = self.convert_to_mapping_data(content, info, download_path)
                            print('保存開始', file=sys.stderr, end=' ... ')
                            self.save_to_cache(map_data, file_cache)
                            logger.info(info['dataset'] + ' 取得完了')
                            print('取得完了', file=sys.stderr)
        except Exception as e:
            logger.exception(e)
            rc = 99
        # geocode検索結果キャッシュを保存
        if self.__geocode_cache_update:
            cache_json = json.dumps(self.__geocode_cache, ensure_ascii=False).replace('"OK"}, "', '"OK"},\n"')
            with open(self.__geocode_cache_path, 'w') as h_cache:
                h_cache.write(cache_json)
        # 復帰
        logger.debug('make_cache_data() ended, rc=' + str(rc))
        return rc

    def url_get(self, url, file, cache=True, dir=None):
        """
        指定されたURLの内容を取得してdownload_dirに保存する
        :param url: str型、取得する資源のURL
        :param file: str型、URLのファイル名部分
        :param cache: boolean型、True=キャッシュを使用する、False:キャッシュを使用しない
        :return: str型、取得したコンテンツ、取得できない場合はNoneを返却する
        :raise: Exception
        """
        logger = logging.getLogger(__name__)
        logger.debug('url_get() start, url=' + url + ', file=' + str(file))
        # 実行
        content = None
        cache_dir = dir
        if cache_dir == None:
            cache_dir = self.download_dir
        if cache \
        and os.path.exists(os.path.join(cache_dir, file)):
            # キャッシュから読み込む
            with open(os.path.join(cache_dir, file), 'rb') as hf:
                content = hf.read()
            logger.debug('Hit cache, file=' + os.path.join(cache_dir, file))
        else:
            # HTTP GET
            time.sleep(1)
            for try_i in range(self.__requests_retry_max):
                try_ok = False
                try:
                    res = requests.get(url, headers=self.__request_headers, proxies=self.__proxies)
                    try_ok = True
                except TimeoutError as e:
                    res = None
                    break
                except requests.exceptions.ProxyError as e:
                    logger.exception(e)
                    res = None
                    if (try_i+1) < self.__requests_retry_max:
                        time.sleep(self.__requests_retry_seconds)
                except Exception as e:
                    logger.exception(e)
                    res = None
                    if (try_i+1) < self.__requests_retry_max:
                        time.sleep(self.__requests_retry_seconds)
                if try_ok:
                    break
            if res is None or res.status_code != requests.codes.ok:
                logger.error('requests error(2): ' + str(res))
            else:
                logger.debug('response.encoding: ' + str(res.encoding))
                content = res.content
            if res is not None:
                res.close()
            # キャッシュディレクイトリに保存する
            if cache and content is not None:
                with open(os.path.join(cache_dir, file), 'wb') as f:
                    f.write(content)
            # logger.debug('OK requests.get(), size=' + str(len(content)))

        # 復帰
        logger.debug('url_get() ended, file=' + str(file))
        return content

    def convert_to_utf8(self, content, rows=0):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('convert_to_utf8() start, rows=' + str(rows))
        # 実行
        data = None
        cont = content
        if rows > 0:
            lines = []
            index = 0
            while len(lines) < (rows+15):
                try:
                    n_index = content[index:].index(b'\n')
                    lines.append(content[index:index+n_index+1])
                    index += n_index + 1
                except Exception as e:
                    lines.append(content[index:])
                    break
            cont = lines[0]
            for i in range(1, len(lines)):
                cont += lines[i]
        res = {'encoding': 'utf-8', 'confidence': 0.5, 'language': 'Japanese'}
        try:
            # ToDo: res = chardet.detect(cont)
            res = detect(cont)
            # res = {'encoding': 'SHIFT_JIS', 'confidence': 0.8719851576994434, 'language': 'Japanese'}
        except Exception as e:
            logger.exception(e)
        logger.debug('detect=' + str(res))
        if res['encoding'] is None \
        or (len(res['encoding']) >= 5 and res['encoding'].upper()[:5] != 'UTF-8'):
            res['encoding'] = 'CP932'
        """ ToDo:
        if res['encoding'] is not None:
            res['encoding'] = res['encoding'].upper()
        if res['encoding'] is None or res['encoding'] == 'ASCII' \
        or res['encoding'] == 'SHIFT_JIS' or res['encoding'] == 'CP932':
            res['encoding'] = 'CP932'
        elif res['encoding'] != 'UTF-8':
            res['encoding'] = 'CP932'
        """
        try:
            data = cont.decode(res['encoding'])
        except Exception as e:
            logger.exception(e)
            data = str(cont)
        # 復帰
        logger.debug('convert_to_utf8() ended.')
        return data

    def convert_to_mapping_data(self, content, info, file):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('convert_to_mapping_data() start.')
        data = None
        if info['format'] == 'CSV':
            data = self.data_from_csv(content, info);
        elif info['format'] in ['TEXT', 'TXT', 'TSV']:
            data = self.data_from_tsv(content, info);
        elif info['format'] == 'XLS':
            data = self.data_from_xls(content, info, file);
        elif info['format'] == 'XLSX':
            data = self.data_from_xlsx(content, info, file);
        # 復帰
        logger.debug('convert_to_mapping_data() ended.')
        return data

    def data_from_csv(self, content, info):
        """
        CSV形式のデータをJSON形式に変換する。
        :param content: str型、CSV形式文字列
        :param info: dict型、形式情報
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.debug('data_from_csv() start.')
        # 実行
        table = self.table_from_csv(content)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.debug('data_from_csv() ended.')
        return map_data

    def data_from_tsv(self, content, info):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('data_from_tsv() start.')
        # 実行
        table = self.table_from_tsv(content)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.debug('data_from_tsv() ended.')
        return map_data

    def data_from_xls(self, content, info, file):
        logger = logging.getLogger(__name__)
        logger.debug('data_from_xls() start.')
        # 実行
        table = self.table_from_xls(file)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.debug('data_from_xls() ended.')
        return map_data

    def data_from_xlsx(self, content, info, file):
        logger = logging.getLogger(__name__)
        logger.debug('data_from_xlsx() start.')
        # 実行
        table = self.table_from_xlsx(file)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.debug('data_from_xlsx() ended.')
        return map_data

    def table_to_mapdata(self, data, info):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('table_to_mapdata() start')
        # 実行
        map_data = []
        # 名称リスト作成
        self.__data_name_dict = {}
        for i, rec in enumerate(data):
            if len(rec) == 0 or (type(rec[0]) == str and (len(rec[0]) == 0 or rec[0][0] == '#')):
                continue
            if i >= info['header'] and len(rec) > 0 \
            and (type(rec[0]) != str or (len(rec[0]) > 0 and rec[0][0] != '#')):
                name = ''
                for n in info['name']:
                    if type(n) == int:
                        if len(rec) > n:
                            name += str(rec[n])
                    else:
                        name += str(n)
                if name in self.__data_name_dict:
                    self.__data_name_dict[name] += 1
                else:
                    self.__data_name_dict[name] = 1
        # テーブルデータをマップ情報に変換する
        headers = data[info['header']-1]
        no = 0;
        for i in range(len(data)):
            if i < info['header']:
                continue;
            if len(data[i]) == 0 or (type(data[i][0]) == str and (len(data[i][0]) == 0 or data[i][0][0] == '#')):
                continue;
            name = ''
            for n in info['name']:
                if type(n) == int:
                    if len(data[i]) > n:
                        name += str(data[i][n])
                else:
                    name += str(n)
            info_items = []
            for j in info['info']:
                if len(data[i]) <= j:
                    continue
                if data[i][j] == '':
                    continue
                info_items.append({headers[j]: data[i][j]})
            # info_items.append({'データセットURL': info['url']})
            error = ''
            no += 1
            id_value = ('000' + str(no))[-4:]
            if info['id'] >= 0:
                if data[i][info['id']] != '':
                    id_value = data[i][info['id']]
            else:
                error += 'id未設定。'
            loc_name = self.state + self.name
            if self.state == self.name:
                loc_name = self.name
            address = ''
            if info['address'] >= 0 and len(data[i]) > info['address']:
                address = data[i][info['address']]
            (lat, lng, msg) = self.lat_lng_from_data(data[i], info, address)
            data_value = {
                    "id": id_value,
                    "locality_code": self.code,
                    "locality_name": loc_name,
                    "kind": info['kind'],
                    "dataset": info['dataset'],
                    "label": name,
                    "lat": lat,
                    "lng": lng,
                    "info": info_items,
            }
            if msg != '':
                data_value["error"] = msg
            map_data.append(data_value)
        self.__data_name_dict = {}
        # 復帰
        logger.debug('table_to_mapdata() ended.')
        return map_data

    def exists_in_cache(self, file, dir=None):
        """
        指定されたファイルがキャッシュに存在するか確認する
        :param file: str型、ファイル名
        :return: boolean型、True=存在する、False=存在しない
        """
        logger = logging.getLogger(__name__)
        logger.debug('exists_in_cache() start, file=' + str(file))
        # 実行
        rc = False
        cache_dir = dir
        if cache_dir is None:
            cache_dir = self.cache_dir
        if os.path.exists(os.path.join(cache_dir, file)):
            rc = True
        # 復帰
        logger.debug('exists_in_cache() ended, rc=' + str(rc))
        return rc

    def save_to_cache(self, mapping_data, file, dir=None):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('save_to_cache() start, file=' + file)
        # 実行
        try:
            cache_dir = dir
            if cache_dir is None:
                cache_dir = self.cache_dir
            if not os.path.exists(cache_dir):
                os.mkdir(cache_dir)
            content = json.dumps(mapping_data, ensure_ascii=False).replace('}, {', '},\n{')
            with open(os.path.join(cache_dir, file), 'w') as f:
                f.write(content)
        except Exception as e:
            logger.exception(e)
            logger.error('path=' + str(os.path.join(cache_dir, file)))
            logger.error('mapping_data=' + str(mapping_data))
            raise Exception('error in store_to_download_dir(), file=' + file)
        # 復帰
        logger.debug('save_to_cache() ended, file=' + str(file))
        return content

    def __save_to_download_dir(self, content, file):
        """
        contentをdownload_dirにファイルとして保存する
        :param content: byte型、ファイルの内容
        :param file: str型、ファイル名
        :return: boolean型、True=正常終了
        :raise Exception: 異常終了
        """
        logger = logging.getLogger(__name__)
        logger.debug('__save_to_download_dir() start, file=' + file)
        logger.debug('save file=' + os.path.join(self.download_dir, file))
        # 実行
        try:
            with open(os.path.join(self.download_dir, file), 'wb') as f:
                f.write(content)
        except Exception as e:
            logger.exception(e)
            raise Exception('error in store_to_download_dir().')
        # 復帰
        logger.debug('__save_to_download_dir() ended.')
        return True

    def lat_lng_from_data(self, data, info, address):
        """
        レコードの緯度・経度を浮動小数点形式で返却する。
        緯度・経度が設定されていない場合は、名称または住所から緯度・経度を求める。
        :param data: Array型、データレコード
        :param info: dict型、データ定義
        :param address: str型、住所
        :return: (緯度, 経度, メッセージ)、緯度=float型、経度=float型、メッセージ=str型
        """
        logger = logging.getLogger(__name__)
        # logger.debug('lat_lng_from_data() start.')
        # logger.debug('lat_lng_from_data() start, data=' + str(data) + ', info=' + str(info))
        # 実行
        lat = ''
        lng = ''
        msg = ''
        name = ''
        for n in info['name']:
            if type(n) == int:
                name += str(data[n])
            else:
                name += str(n)
        if name in self.__data_name_dict \
        and self.__data_name_dict[name] > 1:
            msg += '名称が重複。'
            logger.debug('名称が重複, name=' + name + ', count=' + str(self.__data_name_dict[name]))
        if info['lat'] == -1 or info['lng'] == -1:
            msg += '緯度・経度が未定義。'
        elif len(data) <= info['lat'] or len(data) <= info['lng']:
            msg += '緯度・経度の値無し。'
        elif data[info['lat']] == '' or data[info['lng']] == '':
            msg += '緯度・経度が未設定。'
        else:
            lat = data[info['lat']]
            lng = data[info['lng']]

        if lat == '' or lng == '':
            # 住所から緯度・経度を取得する
            if address == '':
                msg += '住所未設定。'
            else:
                param_addr = address
                if address[:len(self.state)] == self.state:
                    if address[len(self.state):len(self.state)+len(self.name)] == self.name:
                        # 住所に都道府県名と市区町村名が含まれている→加工不要
                        pass
                    else:
                        # 住所に都道府県名が含まれているが市区町村名は含まれていない→市区町村名を挿入する
                        if self.state == self.name:
                            # 都道ぬ件名が重複指定された場合→加工不要
                            pass
                        else:
                            param_addr = self.state + self.name + address[len(self.state):]
                else:
                    if address[:len(self.name)] == self.name:
                        # 住所に市区町村名が含まれている→都道府県名を追加する
                        param_addr = self.state + address
                    else:
                        # 住所に都道府県名も市区町村名も含まれていない→都道府県名と市区町村名を追加する
                        if self.state == self.name:
                            # 都道ぬ件名が重複指定された場合→都道府県名のみを追加する
                            param_addr = self.state + address
                        else:
                            param_addr = self.state + self.name + address
                if param_addr in self.__geocode_cache:
                    result = self.__geocode_cache[param_addr]
                    logger.debug('Hit in __geocode_cache, 住所=' + param_addr)
                else:
                    if True:	# ToDo:
                        result = {'status': 'NG', 'note': '住所から緯度・経度への変換は中止（Google Mapsの課金が発生するため）'}
                    else:
                        geo_url = self.__geocode_url.format(address=param_addr, key=self.__google_maps_api_key)
                        content = self.url_get(geo_url, None, cache=False)
                        if content is None:
                            result = {'status': 'NG'}
                        else:
                            result = json.loads(content)
                            self.__geocode_cache[param_addr] = result
                            self.__geocode_cache_update = True
                logger.debug(param_addr + '：' + str(result))
                if result['status'] == 'OK':
                    if len(result['results']) == 1:
                        res = result['results'][0]
                        if 'location_type' in res['geometry']:
                            # if res['geometry']['location_type'] \
                            #       in ['ROOFTOP', 'RANGE_INTERPOLATED', 'GEOMETRIC_CENTER', 'APPROXIMATE']:
                            if self.state + self.name in res['formatted_address']:
                                lat = res['geometry']['location']['lat']
                                lng = res['geometry']['location']['lng']
                                if res['geometry']['location_type'] == 'ROOFTOP':
                                    msg += '住所から緯度・経度を取得(' + res['geometry']['location_type'] + ')。'
                                else:
                                    msg += '住所から近隣の緯度・経度を取得(' + res['geometry']['location_type']
                                    msg += ')=' + res['formatted_address'] + '。'
                        else:
                            msg += '住所：' + param_addr + '：NG geometry無し。'
                    else:
                        msg += '住所：' + param_addr + '：NG 複数results。'
                else:
                    msg += '住所：' + param_addr + '：NG status=' + result['status'] + '。'
        if (lat == '' or lng == '') and address != '':
            msg += '住所から取得できず。'

        if lat == '' or lng == '':
            # 名称から緯度・経度を取得する
            if self.state == self.name:
                param_name = self.state + ' ' + name
            else:
                param_name = self.state + self.name + ' ' + name
            if param_name in self.__geocode_cache:
                result = self.__geocode_cache[param_name]
                logger.debug('Hit in __geocode_cache, 名称=' + param_name)
            else:
                if True:	# ToDo:
                    result = {'status': 'NG', 'note': '名称から緯度・経度への変換は中止（Google Mapsの課金が発生するため）'}
                else:
                    geo_url = self.__geocode_url.format(address=param_name, key=self.__google_maps_api_key)
                    content = self.url_get(geo_url, None, cache=False)
                    if content is None:
                        result = {'status': 'NG'}
                    else:
                        result = json.loads(content)
                        self.__geocode_cache[param_name] = result
                        self.__geocode_cache_update = True
            logger.debug(param_name + '：' + str(result))
            if result['status'] == 'OK':
                for res in result['results']:
                    if 'geometry' in res:
                        if 'location_type' in res['geometry']:
                            if self.state + self.name not in res['formatted_address']:
                                continue
                            if res['geometry']['location_type'] == 'ROOFTOP':
                                lat = res['geometry']['location']['lat']
                                lng = res['geometry']['location']['lng']
                                msg += '名称から緯度・経度を取得1(' + res['geometry']['location_type']
                                if unicodedata.normalize('NFKC', res['address_components'][0]['long_name']) \
                                         == unicodedata.normalize('NFKC', name):
                                    msg += ')。'
                                else:
                                    msg += ')=' + res['formatted_address'] + '。'
                                    break
                            elif res['geometry']['location_type'] == 'GEOMETRIC_CENTER':
                                if res['address_components'][0]['long_name'] == name:
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称から緯度・経度を取得2(' + res['geometry']['location_type'] + ')。'
                                    break
                                elif len(result['results']) == 1 \
                                and (res['address_components'][0]['long_name'][:2] == name[:2] \
                                  or res['address_components'][0]['long_name'][-2:] == name[-2:]):
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称（類似）から緯度・経度を取得3(' + res['geometry']['location_type']
                                    if unicodedata.normalize('NFKC', res['address_components'][0]['long_name']) \
                                             == unicodedata.normalize('NFKC', name):
                                        msg += ')。'
                                    else:
                                        msg += ')=' + res['formatted_address'] + '。'
                                    # ToDo: msg += '住所未設定。'
                                    break
                            elif address != '':
                                a_addr = unicodedata.normalize('NFKC', address.replace('番地', '−'))
                                # logger.debug('a_addr=' + a_addr)
                                a_match = re.match('[^0-9０-９]*([0-9０-９\-−－ー]*)$', a_addr)
                                if a_match is not None and a_match.group(1) != '':
                                    # logger.debug(str(a_match.groups()))
                                    r_addr = unicodedata.normalize('NFKC', res['formatted_address'].replace('番地', '−'))
                                    # logger.debug('r_addr=' + r_addr)
                                    r_match = re.match('[^0-9０-９]*([0-9０-９\-−－ー]*)$', r_addr)
                                    if r_match is not None and r_match.group(1) != '':
                                        if a_match.group(1) == r_match.group(1):
                                            lat = res['geometry']['location']['lat']
                                            lng = res['geometry']['location']['lng']
                                            msg += '名称から緯度・経度を取得4(' + res['geometry']['location_type'] + ')'
                                            msg += '=番地が同一(' + a_match.group(1) + ')。'
                                            break
                            elif address == '' and len(result['results']) == 1:
                                if res['formatted_address'].find(self.state) >= 0 \
                                and res['formatted_address'].find(self.name) >= 0:
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称から近隣の緯度・経度を取得5(' + res['geometry']['location_type'] \
                                         + ')=' + res['formatted_address'] + '。'
                                    # ToDo: msg += '住所未設定。'
                                    break
                            else:
                                msg += '名称：' + param_name + '：NG location_type=' + res['geometry']['location_type'] + '。'
                        else:
                            msg += '名称：' + param_name + '：NG location_type無し。'
                    else:
                        msg += '名称：' + param_name + '：NG geometry無し。'
            else:
                msg += '名称：' + param_name + '：NG status=' + result['status'] + '。'

        # 浮動小数点形式に変換する
        if type(lat) != float:
            (lat, m) = self.string_to_float(lat)
            if m != '':
                msg += '緯度：' + m
        if type(lng) != float:
            (lng, m) = self.string_to_float(lng)
            if m != '':
                msg += '経度：' + m
        # 復帰
        # logger.debug('lat_lng_from_data() ended, lat=' + str(lat) + ', lng=' + str(lng) + ', msg=' + msg)
        return (lat, lng, msg)

    def string_to_float(self, v_str):
        """
        文字列形式の緯度・経度を浮動小数点に変換する
        :param v_str: str型、文字列形式の緯度または経度
        :return: 浮動小数点形式の緯度または経度。ただし、文字列形式の値が空文字の場合は空文字を返却する
        """
        msg = ''
        msg_correct = ''
        v_float = ''
        try:
            if v_str is None:
                raise Exception('msg:値なし。')
            if v_str == '' or v_str == ' ' or v_str == '　' or v_str == '-' or v_str == '－':
                raise Exception('msg:値が空。')
            v_str = v_str.replace(' ', '').replace(',', '')
            match = RE_FLOAT_FORMAT.search(v_str)
            if match is not None:
                v_str = match.group(1)
            v_strs = v_str.split('.')
            if len(v_strs) == 2:
                # 正しい形式
                v_float = float(v_str)
            elif len(v_strs) == 1:
                # 「.」を入れ忘れた？
                if v_strs[0][0] == '1':
                    v_float = float(v_strs[0][:3] + '.' + v_strs[0][3:])
                else:
                    v_float = float(v_strs[0][:2] + '.' + v_strs[0][2:])
                msg += '度に小数点無し。'
            else:
                # 度.分.秒形式？、日本測地系から世界測地系に簡易補正
                v_float = float(v_strs[0]) \
                        + float(v_strs[1]) / 60.0
                msg += '度.分.秒形式から変換'
                if len(v_strs) >= 3:
                    if len(v_strs) == 3:
                        v_float += float(v_strs[2]) / 3600.0
                    else:
                        v_float += float(v_strs[2]+'.'+v_strs[3]) / 3600.0
                    # 沼津市地点での簡易補正（秒の指定が無ければ誤差の範囲とみなす）
                    if v_float > 90.0:
                        v_float -= 11.29 / 3600
                        msg_correct = '-11.29秒'
                    else:
                        v_float += 11.85 / 3600
                        msg_correct = '+11.85秒'
                    msg += '・簡易補正'
                msg += '(' + v_str + msg_correct + ')。'
        except Exception as e:
            if len(e.args) > 0 and type(e.args[0]) == str and e.args[0][:4] == 'msg:':
                msg += e.args[0][4:]
            else:
                logger = logging.getLogger(__name__)
                logger.exception(e)
                msg += str(e)
        return (v_float, msg)

    def download_datasets_in_site(self, name):
        """
        オープンデータカタログサイトから緯度・経度または住所を含むデータセットをダウンロードする。
        :param state: str型、オープンデータカタログサイトの自治体名：都道府県名
        :param url: str型、自治体名オープンデータカタログサイトのurl：package_list
        :return: 
        """
        logger = logging.getLogger(__name__)
        logger.debug('download_datasets_in_site() start, name=' + str(name))
        # 実行
        rc = 0
        for site in OPENDATA_SITES:
            if name is not None and site['name'] != name:
                # 対象外サイト
                continue
            self.state = site['name']
            print('サイト=' + site['code'] + site['name'] + '：' + site['api_url'], file=sys.stderr)
            logger.debug('site=' + str(site))
            # パッケージリストを取得する
            packages_dir = os.path.join(self.download_dir, 
                    MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
            jpackages = self.get_package_list(site)
            # geocode検索結果キャッシュを初期化
            self.__geocode_cache = {}
            self.__geocode_cache_update = False
            self.__geocode_cache_path = os.path.join(packages_dir, self.__geocode_cache_file.format(site['name']))
            if os.path.exists(self.__geocode_cache_path):
                with open(self.__geocode_cache_path, 'r') as hf:
                    self.__geocode_cache = json.loads(hf.read())
            # キャッシュディレクトリを確認する
            cache_site_dir = os.path.join(self.cache_dir, \
                    MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
            if not os.path.exists(cache_site_dir):
                os.mkdir(cache_site_dir)
            # 対象地方自治体コード表を取得する
            self.locality_dict = {}
            with open(os.path.join(self.APP_DIR, 'crawler', LOCALITY_CODE_FILE), 'r') as hf:
                hcsv = csv.reader(hf)
                self.locality_dict = {rec[2]:rec[0] for rec in hcsv if rec[1] == site['name']}
            info_list = {}
            info_list_file = os.path.join(packages_dir, '.info_list.json')
            info_list_update = False
            if os.path.exists(info_list_file):
                # データセットのマップ情報を読み込む
                with open(info_list_file, 'r') as hf:
                    info_list = json.loads(hf.read())
            # 全パッケージを処理する
            # ToDo: jpackages['result'] = ['t131156d0000000017']
            for pkg_i, package in enumerate(jpackages['result']):
                try:
                    logger.debug('package=' + package)
                    # パッケージファイルを取得する
                    jpackage = self.get_package_json(package, site)
                    if 'result' not in jpackage:
                        print(str(pkg_i+1) + '/' + str(len(jpackages['result'])) \
                                + '：不正なパッケージ', file=sys.stderr)
                        logger.debug('不正なパッケージ')
                        continue
                    locality_name, locality_code = self.get_locality_from_package(jpackage, site['code'])
                    self.name = locality_name
                    self.code = locality_code
                    print(str(pkg_i+1) + '/' + str(len(jpackages['result'])) + '：' \
                            + locality_code + locality_name + '：', file=sys.stderr, end='')
                    # データセットのマップ情報を確認する
                    dataset_name = jpackage['result']['name']
                    if 'title' in jpackage['result']:
                        if len(jpackage['result']['title']) <= 128:
                            dataset_name = jpackage['result']['title']
                    print(dataset_name + '：', file=sys.stderr, end='')
                    if any([dataset_name.find(w)>=0 for w in IGNORE_WORD_IN_DATASET_NAME]):
                        print('特定単語を含むデータセットは除外', file=sys.stderr)
                        logger.debug('特定単語を含むデータセットは除外')
                        continue
                    info_key = self.state + '_' + self.name + '_' + dataset_name
                    if info_key not in info_list or info_list[info_key] == {}:
                        # データセットのマップ情報を作成する
                        info = self.make_map_info(dataset_name, jpackage, site)
                        if info_key not in info_list \
                        or info_list[info_key] != info:
                            info_list[info_key] = info
                            info_list_update = True
                    info = info_list[info_key]
                    logger.debug(info_key + ' info('+str(len(info))+')=' + str(info))
                    if info == []:
                        print('マップ情報不足', file=sys.stderr)
                        logger.debug('マップ情報不足')
                        continue

                    # リソース毎に処理する
                    for res_i, res_def in enumerate(info):
                        # マップデータのキャッシュを確認する
                        cache_dir = os.path.join(cache_site_dir, \
                                MY_CONFIG['dir_name'].format(code=locality_code, name=locality_name))
                        file_cache = str(pathlib.Path(res_def['file']).with_suffix('.json'))
                        if self.exists_in_cache(file_cache, dir=cache_dir):
                            # キャッシュに存在する
                            print('[' + str(res_i) + ']マップデータがキャッシュに存在する', file=sys.stderr)
                            continue
                        # データーセットを取得する
                        dataset_dir = os.path.join(packages_dir, \
                                MY_CONFIG['dir_name'].format(code=locality_code, name=locality_name))
                        if not os.path.exists(dataset_dir):
                            os.mkdir(dataset_dir)
                        dataset_path = os.path.join(dataset_dir, res_def['file'])
                        content = self.url_get(res_def['url'], res_def['file'], dir=dataset_dir)
                        if content is None:
                            print('[' + str(res_i) + ']データセット取得失敗', file=sys.stderr)
                            continue
                        print('[' + str(res_i) + ']データセット取得済：', file=sys.stderr, end='')
                        # 必要ならコード変換を行う
                        if res_def['format'] in ['CSV', 'TEXT', 'TXT', 'TSV']:
                           content = self.convert_to_utf8(content)
                           if content is None:
                               print('[' + str(res_i) + ']形式誤り', file=sys.stderr)
                               continue
                        # 表形式をマップデータ(JSON形式)に変換する
                        map_data = self.content_to_mapping_data(content, res_def, dataset_path)
                        print('[' + str(res_i) + ']マップデータ作成済：', file=sys.stderr, end='')
                        # マップデータをキャッシュに保存する
                        self.save_to_cache(map_data, file_cache, dir=cache_dir)
                        print('[' + str(res_i) + ']キャッシュ済', file=sys.stderr)
                        logger.info('[' + str(res_i) + ']保存完了：')

                except Exception as e:
                    if len(e.args) > 0 and type(e.args[0]) == str and e.args[0][:4] == 'msg:':
                        print(e.args[0][4:], file=sys.stderr)
                    else:
                        logger.exception(e)
                        print('', file=sys.stderr)
                        rc = 99
                        break

            # データセットのマップ情報を保存する
            if info_list_update:
                with open(info_list_file, 'w') as hf:
                    hf.write(json.dumps(info_list, ensure_ascii=False).replace('], "', '],\n"'))
            # geocode検索結果キャッシュを保存
            if self.__geocode_cache_update:
                cache_json = json.dumps(self.__geocode_cache, ensure_ascii=False).replace('"OK"}, "', '"OK"},\n"')
                with open(self.__geocode_cache_path, 'w') as h_cache:
                    h_cache.write(cache_json)

        # 復帰
        logger.debug('download_datasets_in_site() ended')
        return rc

    def get_locality_from_package(self, jpackage, default_code):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('get_locality_from_package() start')
        locality_name = ''
        locality_code = ''
        if ('groups' not in jpackage['result'] \
          or len(jpackage['result']['groups']) < 1) \
        and ('areas' not in jpackage['result'] \
          or len(jpackage['result']['areas']) < 1) \
        and ('organization' not in jpackage['result'] \
          or len(jpackage['result']['organization']) < 1):
            locality_name = self.state
        else:
            if 'groups' in jpackage['result'] \
            and len(jpackage['result']['groups']) > 0:
                i_group = 0
                if len(jpackage['result']['groups']) > 1:
                    if jpackage['result']['groups'][0]['id'] == 1:
                        i_group = 1
                if 'trailing_name' in jpackage['result']['groups'][i_group]:
                    locality_name = jpackage['result']['groups'][i_group]['trailing_name']
                    if locality_name == jpackage['result']['groups'][i_group]['name'].split('/')[0]:
                        locality_name = jpackage['result']['areas'][0]['name']
            if locality_name == '' \
            and 'areas' in jpackage['result']:
                locality_name = jpackage['result']['areas'][0]['name']
            if locality_name == '' \
            and 'organization' in jpackage['result']:
                if re.match('^[0-9a-zA-Z]+$', jpackage['result']['organization']['name']) == None:
                    locality_name = jpackage['result']['organization']['name']
                elif jpackage['result']['organization']['title'][-1] in ['市','区','町','村']:
                    if jpackage['result']['organization']['title'][:len(self.state)] == self.state:
                        locality_name = jpackage['result']['organization']['title'][len(self.state):]
                    else:
                        locality_name = jpackage['result']['organization']['title']
        if locality_name == '':
            locality_name = self.state
        if locality_name not in self.locality_dict:
            locality_code = default_code
        else:
            locality_code = self.locality_dict[locality_name]
        logger.debug('get_locality_from_package() ended, locality_name=' + locality_name + ', locality_code=' + locality_code)
        return locality_name, locality_code

    def get_package_list(self, site):
        # パッケージリストを取得する
        logger = logging.getLogger(__name__)
        logger.debug('get_package_list() start, site=' + str(site))
        jpackages = []
        packages_dir = os.path.join(self.download_dir, 
                MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
        if not os.path.exists(packages_dir):
            os.mkdir(packages_dir)
            logger.info('make site dir=' + packages_dir)
        packages_file = 'package_list.json'
        url = site['api_url'] + 'package_list'
        content = self.url_get(url, packages_file, dir=packages_dir)
        if content is not None:
            jpackages = json.loads(content)
        # 復帰
        logger.debug('get_package_list() ended')
        return jpackages

    def get_package_json(self, package, site):
        # パッケージファイルを取得する
        logger = logging.getLogger(__name__)
        logger.debug('get_package_json() start, package=' + str(package))
        # 実行
        jpackage = {}
        packages_dir = os.path.join(self.download_dir, 
                MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
        package_file = package+'.package'
        url = site['api_url'] + 'package_show?id=' + package
        content = self.url_get(url, package_file, dir=packages_dir)
        if content is not None:
            jpackage = json.loads(content)
            logger.debug('jpackage=' + str(jpackage))
            if not jpackage['success']:
                logger.error('ERROR: error in result, result=' + str(jpackage))
                raise Exception('msg:パッケージ取得失敗：' + str(jpackage))
            if jpackage['result']['type'] != 'dataset':
                logger.error('ERROR: error in dataset type, type=' + jpackage['result']['type'])
                raise Exception('msg:datasetを含まないパッケージ：' + str(jpackage))
        # 復帰
        logger.debug('get_package_json() ended')
        return jpackage

    def make_map_info(self, dataset_name, jpackage, site):
       # マップ情報を作成する
        logger = logging.getLogger(__name__)
        logger.debug('make_map_info() start, dataset_name=' + str(dataset_name))
        # 実行
        info = []
        packages_dir = os.path.join(self.download_dir, 
                MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
        # データセットの所有者
        logger.debug('自治体：code=' + self.code + ', name=' + self.name)
        # 処理対象のリソースを抽出する
        rNos = self.get_unique_resources(jpackage)
        if len(rNos) == 0:
            return info
        kind = self.get_kind_normalized(dataset_name)
        if kind is None:
            logger.debug('kind is not found.')
            return info
        # リソース毎に処理する
        for rNo in rNos:
            resource = jpackage['result']['resources'][rNo]
            logger.debug('resource=' + str(resource))
            # データセットを取得する
            # パッケージのリソース内容確認
            dataset_file = ''
            if 'filename' in resource:
                if resource['filename'] is not None and resource['filename'] != '':
                    dataset_file = resource['filename']
            if dataset_file == '' and 'download_url' in resource:
                if resource['download_url'] is not None and resource['download_url'] != '':
                    dataset_file = resource['download_url'].split('/')[-1]
            if dataset_file == '' and 'url' in resource:
                if resource['url'] is not None and resource['url'] != '':
                    dataset_file = resource['url'].split('/')[-1]
            if dataset_file == '':
                dataset_file = resource['name'] + '.' + resource['format'].lower()
            if len(dataset_file.encode('UTF-8')) >= 128:
                dataset_file = dataset_file[:63] + '_' + dataset_file[-63:]
            if dataset_file.find('/') >= 0:
                dataset_file = dataset_file.replace('/', '_')
            if dataset_file.find('\n') >= 0:
                dataset_file = dataset_file.replace('\n', '_')
            if len(dataset_file.split('.')) == 1:
                dataset_file += '.' + resource['format'].lower()
            elif dataset_file.split('.')[-1].upper() != resource['format'].upper():
                dataset_file = '.'.join(dataset_file.split('.')[:-1]) + '.' + resource['format'].lower()
            dataset_dir = os.path.join(packages_dir, \
                    MY_CONFIG['dir_name'].format(code=self.code, name=self.name))
            if not os.path.exists(dataset_dir):
                os.mkdir(dataset_dir)
                logger.info('mkdir locality directory=' + dataset_dir)
            dataset_path = os.path.join(dataset_dir, dataset_file)
            if 'url' not in resource and 'download_url' not in resource:
                logger.debug('url and download_url key not found.')
                continue
            if 'url' in resource:
                url = resource['url']
            if 'download_url' in resource:
                url = resource['download_url']
            # リソース定義の初期化
            res_def = {
                "kind": kind,
                "dataset": dataset_name,
                "url": url,
                "file": dataset_file,
                "format": resource['format'],
                "lat": -1,
                "lng": -1,
                "name": [-1],
                "id": -1,
                "info": [-1],
                "address": -1,
                "header": -1
            }
            # データセットを取得する
            content = self.url_get(url, res_def['file'], dir=dataset_dir)
            if content is None:
                logger.error('content not found, url=' + url)
                continue
            # HTML形式でダウンロードページならダウンロードする
            redirect = self.check_html_page(content, res_def['url'], dataset_dir)
            if redirect is not None:
                res_def['format'] = redirect['format']
                res_def['file'] = redirect['file']
                res_def['url'] = redirect['url']
                content = redirect['content']
                dataset_path = os.path.join(dataset_dir, res_def['file'])
            # コンテンツの内容でformatを検査する
            head_bytes8 = bytes(content[:8])
            if len(content) > 8 and head_bytes8 == bytes.fromhex('d0cf11e0a1b11ae1'):
                # XLS形式：DOC/XLS/PPTも同じだが判定省略
                if res_def['format'].upper() != 'XLS':
                    logger.warning('format変更, ' + res_def['format'] + ' -> XLS')
                    res_def['format'] = 'XLS'
            elif len(content) > 6 and head_bytes8[:6] == bytes.fromhex('504b03041400'):
                # XLSX形式：ZIP/DOCX/XLSX/PPTXも同じだが判定省略
                if res_def['format'].upper() != 'XLSX':
                    logger.warning('format変更, ' + res_def['format'] + ' -> XLSX')
                    res_def['format'] = 'XLSX'
            elif content.count(b'\n') <= content.count(b','):
                # CSV形式
                if res_def['format'].upper() != 'CSV':
                    logger.warning('format変更, ' + res_def['format'] + ' -> CSV')
                    res_def['format'] = 'CSV'
            elif content.count(b'\n') <= content.count(b'\t'):
                # TSV形式：TXT/TEXTも同じ
                if res_def['format'].upper() not in ['TXT', 'TSV', 'TEXT']:
                    logger.warning('format変更, ' + res_def['format'] + ' -> TXT')
                    res_def['format'] = 'TXT'
            else:
                # 指定のまま
                logger.warning('format不明, ' + res_def['format'])
            # Excel形式でファイルの拡張子が違っている場合は変更する
            if res_def['format'] in ['XLSX', 'XLS'] \
            and res_def['file'].split('.')[-1].upper() != res_def['format'].upper():
                dist_name = '.'.join(res_def['file'].split('.')[:-1]) + '.' + res_def['format'].lower()
                os.rename(os.path.join(dataset_dir, res_def['file']),
                        os.path.join(dataset_dir, dist_name))
                res_def['file'] = dist_name
                dataset_path = os.path.join(dataset_dir, res_def['file'])
            # 必要ならコード変換を行う
            if res_def['format'] in ['CSV', 'TEXT', 'TSV', 'TXT']:
                content = self.convert_to_utf8(content, rows=HEADER_ROWS)
                if content is None:
                    logger.error('content format error, format=' + res_def['format'])
                    continue
            # ダウンロードデータを表形式に変換する
            if res_def['format'] == 'CSV':
                # CSV形式
                logger.debug('format is ' + res_def['format'])
                table = self.table_from_csv(content)
            elif res_def['format'] in ['TEXT', 'TXT', 'TSV']:
                # テキスト形式(TSV)
                logger.debug('format is ' + res_def['format'])
                table = self.table_from_tsv(content)
            elif res_def['format'] == 'XLS':
                # Excel形式(XLS)
                logger.debug('format is ' + res_def['format'])
                table = self.table_from_xls(dataset_path, rows=HEADER_ROWS)
            elif res_def['format'] == 'XLSX':
                # Excel形式(XLSX)
                logger.debug('format is ' + res_def['format'])
                table = self.table_from_xlsx(dataset_path, rows=HEADER_ROWS)
            else:
                # 未サポートの形式
                logger.error('ERROR in format, format=' + resource['format'])
                # raise Exception('ToDo: format=' + resource['format'])
                continue
            # データの項目名からマップ情報を完成する
            title_rows = [list(row) for row in table[:min(HEADER_ROWS , len(table))]]
            for i_row in range(len(title_rows)):
                for i_col in range(len(title_rows[i_row])):
                    if title_rows[i_row][i_col] is None:
                        title_rows[i_row][i_col] = ''
            for t in range(len(title_rows)):
                if title_rows[t] == []:
                    continue
                address = [i for i in range(len(title_rows[t])) if title_rows[t][i] in ['住所','所在地']]
                if len(address) == 0:
                    # 住所項目なし
                    address = [-1]
                address = address[0]
                lat = [i for i in range(len(title_rows[t])) if title_rows[t][i] in ['緯度','北緯','lat','LAT'] or (type(title_rows[t][i])==str and len(title_rows[t][i])>4 and title_rows[t][i][-4:]=='#lat')]
                if len(lat) == 0:
                    # 緯度項目なし
                    lat = [-1]
                lat = lat[0]
                lng = [i for i in range(len(title_rows[t])) if title_rows[t][i] in ['経度','東経','lng','LNG'] or (type(title_rows[t][i])==str and len(title_rows[t][i])>5 and title_rows[t][i][-5:]=='#long')]
                if len(lng) == 0:
                    # 経度項目なし
                    lng = [-1]
                lng = lng[0]
                name = [i for i in range(len(title_rows[t])) if title_rows[t][i] in ['名称','施設名','施設名称','場所名','店名','薬局名','駐輪場名','介護サービス事業所名称'] or (type(title_rows[t][i])==str and len(title_rows[t][i])>6 and title_rows[t][i][-6:]=='#label')]
                if len(name) > 0:
                    name = [name[0]]
                else:
                    # 名称項目なし
                    name = [-1]
                    if '種別' in title_rows[t] and '住所' in title_rows[t]:
                        name = [title_rows[t].index('種別'), title_rows[t].index('住所')]
                id = [i for i in range(len(title_rows[t])) if title_rows[t][i] in ['ID','id','NO','no']]
                if len(id) == 0:
                    # ID項目なし
                    id = [-1]
                id = id[0]

                res_def['lat'] = lat
                res_def['lng'] = lng
                res_def['name'] = name
                res_def['id'] = id
                res_def['address'] = address
                res_def['header'] = t + 1
                used_index = [res_def['lat'],res_def['lng'],res_def['id']]
                used_index.extend(res_def['name'])
                res_def['info'] = [i for i in range(len(title_rows[res_def['header']-1])) if i not in used_index]
                if (res_def['lat'] >= 0 and res_def['lng'] >= 0) \
                and res_def['name'] != [-1]:
                    logger.debug('res_def=' + str(res_def))
                    break
                # 必須項目なし
                # raise Exception('ToDo: 必須項目なし')
            # 最終確認
            if (res_def['lat'] < 0 or res_def['lng'] < 0) \
            or res_def['name'] == [-1]:
                # 必須項目なし
                pass
            else:
                logger.debug('headers=' + str(title_rows[res_def['header']-1]))
                info.append(res_def)

        # 復帰
        logger.debug('make_map_info() ended, info=' + str(info))
        return info

    def get_unique_resources(self, jpackage):
        """ リソース群の中から処理対象とするリソースを抽出する
        """
        logger = logging.getLogger(__name__)
        logger.debug('get_unique_resources() start.')
        # 実行
        # 判断材料を抽出する
        names =     [v['name']     if 'name'     in v else None for v in jpackage['result']['resources']]
        formats =   [v['format']   if 'format'   in v else None for v in jpackage['result']['resources']]
        urls =      [v['url']      if 'url'      in v else None for v in jpackage['result']['resources']]
        downloads = [v['download_url'] if 'download_url' in v else None for v in jpackage['result']['resources']]
        urls = [downloads[i] if downloads[i] is not None else urls[i] for i in range(len(urls))]
        urls = [v.split('/')[-1] if v is not None else None for v in urls]
        pkg_resources = [
            {'no': i, 'name': names[i], 'format': formats[i], 'url': urls[i]} for i in range(len(names))
        ]
        # logger.debug('pkg_resources=' + str(pkg_resources).replace('}, {', '},\n{'))
        target_resources = pkg_resources
        # (1)formatが「CSV、TXT、XLSX、XLS、HTML」以外を削除する
        target_resources = [v for v in target_resources if v['format'] in ['CSV','TXT','XLSX','XLS','HTML']]
        # logger.debug('[1]target_resources=' + str(target_resources))
        # (2)１件以下なら終了
        if len(target_resources) <= 1:
            # logger.debug('[2]target_resources=' + str(target_resources))
            rNos = [v['no'] for v in target_resources]
            logger.debug('get_unique_resources() ended, rNos=' + str(rNos))
            return rNos
        # logger.debug('[2]target_resources=' + str(target_resources))
        # (3)「関連ホームページ」を削除する
        target_resources = [v for v in target_resources if not (v['name'] == '関連ホームページ' and v['format'] == 'HTML')]
        # logger.debug('[3]target_resources=' + str(target_resources))
        # (4)nameに拡張子が設定されている場合は削除する
        for i in range(len(target_resources)):
            for s in ['.', '_']:
                name_split = target_resources[i]['name'].split(s)
                # logger.debug('[4]name_split=' + str(name_split))
                if len(name_split) > 1:
                    if name_split[-1].upper() == target_resources[i]['format'].upper():
                        # logger.debug('[5]format SAME!')
                        target_resources[i]['name'] = s.join(name_split[:-1])
        # logger.debug('[6]target_resources=' + str(target_resources))
        # (5)nameが同じでformatが異なるものは、優先順位「CSV>TXT>XLSX>XLS>HTML」で１つにする
        names = [v['name'] for v in target_resources]
        uniq_names = set(names)
        if len(names) > len(uniq_names):
            uniq_resources = []
            uniq_urls = []
            for n in uniq_names:
                f_exists = False
                for f in ['CSV','TXT','XLSX','XLS','HTML']:
                    for i in range(len(target_resources)):
                        if target_resources[i]['name'] != n:
                            continue
                        if target_resources[i]['format'] != f:
                            continue
                        if target_resources[i]['url'] is None:
                            continue
                        if target_resources[i]['url'] in uniq_urls:
                            continue
                        # logger.debug('[7]target_resources[i]=' + str(target_resources[i]))
                        uniq_resources.append(target_resources[i])
                        uniq_urls.append(target_resources[i]['url'])
                        f_exists = True
                    if f_exists:
                        break
            target_resources = uniq_resources
        # logger.debug('[8]target_resources=' + str(target_resources))
        rNos = [v['no'] for v in target_resources]
        # 復帰
        logger.debug('get_unique_resources() ended, rNos=' + str(rNos))
        return rNos

    def check_html_page(self, content, url, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('check_html_page() start.')
        # 実行
        redirect = None
        cont_str = content.decode('UTF-8','ignore').replace('\r\n','\n')
        if len(cont_str) > len(XHTML_PAGE) and cont_str[:len(XHTML_PAGE)] != XHTML_PAGE:
            logger.debug('content is not HTML')
        elif cont_str[:len(LINKDATA_DOWNLOAD_PAGE)] == LINKDATA_DOWNLOAD_PAGE:
            redirect = self.check_html_page_linkdata(cont_str, url, dir)
        else:
            redirect = self.check_html_page_xhtml(cont_str, url, dir)
        # 復帰
        logger.debug('check_html_page() ended.')
        return redirect

    def check_html_page_xhtml(self, cont_str, url, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('check_html_page_xhtml() start.')
        # 実行
        redirect = None
        soup = None
        format = None
        re_url = None
        try:
            soup = BeautifulSoup(cont_str, 'html.parser')
            if soup is None:
                logger.warning('content is not HTML.')
                return redirect
            list_download = soup.select('.list-download')
            if list_download is None or len(list_download) == 0:
                logger.warning('content not have list-download class.')
                return redirect
            downloads = list_download[0].select('.download')
            if downloads is None:
                logger.warning('content not have download class.')
                return redirect
            for download in downloads:
                if 'data-downloadpath' not in download.attrs \
                or download.attrs['data-downloadpath'] == '':
                    logger.warning('content not have data-downloadpath attribute.')
                    continue
                path = download.attrs['data-downloadpath']
                format = None
                text = download.text.upper()
                for f in ['CSV', 'TEXT', 'TXT', 'TSV', 'XLSX', 'XLS']:
                    if text.find(f) < 0:
                        continue
                    format = f
                    break
                if format is None:
                    continue
                re_url = path
                if re_url[0] == '/':
                    re_url = '/'.join(url.split('/')[:3]) + path
                break

        except Exception as e:
            logger.exception(e)
            return redirect
        finally:
            if soup is not None:
                soup.clear()
                soup = None

        if format is not None and re_url is not None:
            redirect = { 'url': re_url, 'file': re_url.split('/')[-1], 'format': format }
            redirect['content'] = self.url_get(redirect['url'], redirect['file'], dir=dir)
            if redirect['content'] is None:
                logger.warning('faild in getting redirect content')
                redirect = None

        # 復帰
        logger.debug('check_html_page_xhtml() ended, redirect=' + str(redirect is not None))
        return redirect

    def check_html_page_linkdata(self, cont_str, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('check_html_page_linkdata() start.')
        TYPE_TO_FORMAT = {
                'Table Data(CSV)': 'CSV',
                'Table Data(Text)': 'TEXT',
                'Table Data(Excel)': 'Excel',
                'RDF (Turtle)': None
        }
        # 実行
        redirect = None
        url_base = LINKDATA_URL_BASE
        match = re.search(r"LD.systemUrl = '([^']+)';", cont_str)
        if match is not None:
            url_base = match.group(1)
            if url_base[-1] == '/':
                url_base = url_base[:-1]
        else:
            logger.debug('check_html_page_linkdata() ended, faild-2')
            # return redirect
        attrs = []
        soup = None
        try:
            soup = BeautifulSoup(cont_str, 'html.parser')
            downloads = soup.select('.downloadFileInfoContainer')
            for download in downloads:
                type = download.select('.downloadFileNameAndType')[0].find('div').text.replace('\n','').replace('\t','')
                version = download.select('.downloadFileVersion')[0].find('div').text.replace('\n','').replace('\t','')
                size = download.select('.downloadFileSize')[0].find('div').text.replace('\n','').replace('\t','')
                url = url_base + download.select('.downloadFileDownload')[0].find('a').attrs['href']
                file = url[url.rfind('/')+1:]
                if file.find('/') >= 0:
                    file = file.replace('/', '_')
                if type in TYPE_TO_FORMAT:
                    format = TYPE_TO_FORMAT[type]
                    if format == 'Excel':
                        if file[-4:].upper() == '.XLS':
                            format = 'XLS'
                        elif file[-5:].upper() == '.XLSX':
                            format = 'XLSX'
                        else:
                            logger.debug('check_html_page_linkdata() ended, faild-3')
                            return redirect
                attrs.append({
                    'format': format,
                    'version': version,
                    'size': size,
                    'url': url,
                    'file': file
                })

        except Exception as e:
            logger.exception(e)
            logger.debug('check_html_page_linkdata() ended, faild-4')
            return redirect
        finally:
            if soup is not None:
                soup.clear()
                soup = None
        logger.debug('attrs=' + str(attrs))

        if len(attrs) == 0:
            logger.debug('check_html_page_linkdata() ended, faild-5')
            return redirect
        for format in ['CSV', 'TEXT', 'TXT', 'TSV', 'XLSX', 'XLS']:
            for attr in attrs:
                if attr['format'] == format:
                    redirect = attr
                    break
            if redirect is not None:
                break
        if redirect is None:
            logger.debug('check_html_page_linkdata() ended, faild-5')
            return redirect
        # リダイレクト先をダウンロードする
        redirect['content'] = self.url_get(redirect['url'], redirect['file'], dir=dir)
        if redirect['content'] is None:
            logger.debug('check_html_page_linkdata() ended, faild()')
            redirect = None
        # 復帰
        logger.debug('check_html_page_linkdata() ended, format=' + redirect['format'])
        return redirect

    def get_kind_normalized(self, kind):
        # 正規化種別リストに従って種別を決定する
        n_kind = None
        for key, value in KIND_LIST_NORMALIZED.items():
            match = value.search(kind)
            if match is None:
                continue
            n_kind = key
            break
        return n_kind

    def content_to_mapping_data(self, content, info, file):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('content_to_mapping_data() start.')
        data = None
        if info['format'] == 'CSV':
            data = self.data_from_csv(content, info);
        elif info['format'] in ['TEXT', 'TXT', 'TSV']:
            data = self.data_from_tsv(content, info);
        elif info['format'] == 'XLS':
            data = self.data_from_xls(content, info, file);
        elif info['format'] == 'XLSX':
            data = self.data_from_xlsx(content, info, file);
        # 復帰
        logger.debug('content_to_mapping_data() ended.')
        return data

    def table_from_csv(self, content):
        """
        CSV形式のデータをテーブル形式に変換する。
        :param content: str型、CSV形式文字列
        :return: list型、list形式行のlist形式
        """
        logger = logging.getLogger(__name__)
        logger.debug('table_from_csv() start.')
        # 実行
        table = [v for v in csv.reader(content.replace('\r\n','\n').replace('\r','\n').split('\n'))]
        # 復帰
        logger.debug('table_from_csv() ended, table[:3]=' + str(table[:3]))
        return table

    def table_from_tsv(self, content):
        """
        """
        logger = logging.getLogger(__name__)
        logger.debug('table_from_tsv() start.')
        # 実行
        table = [v for v in csv.reader(content.replace('\r\n', '\n').replace('\r','\n').split('\n'),
                delimiter='\t', quotechar=None)]
        if len(table) > 0 and len(table[0]) > 0 \
        and type(table[0][0]) == str and table[0][0] == '#LINK':
            # LinkTable形式
            logger.debug('#LINK形式')
            property = True
            link_data = []
            for i in range(len(table)):
                if len(table[i]) == 0 or len(table[i][0]) == 0:
                    continue;
                if property:
                    if table[i][0][0] == '#':
                        if table[i][0] == '#property':
                            link_data.append(['id'] + table[i][1:])
                        continue
                    else:
                        property = False
                link_data.append(table[i])
            table = link_data
        # 復帰
        logger.debug('table_from_tsv() ended, table[:3]=' + str(table[:3]))
        return table

    def table_from_xls(self, file, rows=0):
        """
        Excel形式(XLS)をテーブル形式に変換する。
        """
        logger = logging.getLogger(__name__)
        logger.debug('table_from_xls() start, rows=' + str(rows))
        # 実行
        table = []
        try:
            excel_book = xlrd.open_workbook(file)
            excel_sheet = excel_book.sheet_by_index(0)
            for row in range(excel_sheet.nrows):
                row_value = []
                for col in range(excel_sheet.ncols):
                    v = excel_sheet.cell(row, col).value
                    if v is None:
                        v = ''
                    elif isinstance(v, datetime.time):
                        v = str(v)
                        if v[-3:] == ':00':
                            v = v[:-3]
                    elif isinstance(v, datetime.datetime):
                        v = str(v)
                        if v[-9:] == ' 00:00:00':
                            v = v[:-9]
                        elif v[-3:] == ':00':
                            v = v[:-3]
                    row_value.append(v)
                table.append(row_value)
                if rows > 0  and len(table) >= rows:
                    break
            excel_book = None
            # logger.debug('table=' + str(table).replace('], [', '],\n['))
            logger.debug('table[:3]=' + str(table[:3]))
        except Exception as e:
            logger.error('error in Excel(XLS) data, file=' + file)
            logger.exception(e)
            table = []
            print('Excel形式(XLS)誤り：' + file, file=sys.stderr, end='')
        # 復帰
        logger.debug('table_from_xls() ended')
        return table

    def table_from_xlsx(self, file, rows=0):
        """
        Excel形式(XLSX)をテーブル形式に変換する。
        """
        logger = logging.getLogger(__name__)
        logger.debug('table_from_xlsx() start, rows=' + str(rows))
        # 実行
        table = []
        try:
            excel_book = openpyxl.load_workbook(file)
            excel_sheet = excel_book.worksheets[0]
            for row in excel_sheet.iter_rows():
                row_value = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        v = ''
                    elif isinstance(v, datetime.time):
                        v = str(v)
                        if v[-3:] == ':00':
                            v = v[:-3]
                    elif isinstance(v, datetime.datetime):
                        v = str(v)
                        if v[-9:] == ' 00:00:00':
                            v = v[:-9]
                        elif v[-3:] == ':00':
                            v = v[:-3]
                    row_value.append(v)
                table.append(row_value)
                if rows > 0  and len(table) >= rows:
                    break
            excel_book.close()
            # logger.debug('table=' + str(table).replace('], [', '],\n['))
            logger.debug('table[:3]=' + str(table[:3]))
        except Exception as e:
            logger.error('error in Excel data(XLSX), file=' + file)
            logger.exception(e)
            table = []
            print('Excel形式(XLSX)誤り：' + file, file=sys.stderr, end='')
        # 復帰
        logger.debug('table_from_xlsx() ended')
        return table

def setup_logger(name, level, log_file='crawler.log', log_dir='log'):
    """
    コマンド実行時のログ初期化
    :param name: str型、関数、__main__
    :param level: int型、logging.INFO、logging.DEBUG、...
    :param log_file: str型/stderr、ログファイル名またはstderr
    :param log_dir: str型、ログディレクトリ名、log_fileがログファイル名の場合に有効なパラメタ
    :return: logger
    """
    logger = logging.getLogger(name)
    logger.parent.setLevel(level)
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    log_handler_format = logging.Formatter(log_format)
    if type(log_file) == str:
        # ファイル出力ハンドラ
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        log_file_handler = logging.FileHandler(os.path.join(log_dir, log_file))
        # log_file_handler= logging.handlers.RotatingFileHandler(os.path.join(log_dir, log_file), maxBytes=1024000, backupCount=5)
        log_file_handler.setLevel(level)
        log_file_handler.setFormatter(log_handler_format)
        logger.addHandler(log_file_handler)
    else:
        # 標準エラー出力ハンドラ
        log_stream_handler = logging.StreamHandler()
        log_stream_handler.setLevel(level)
        log_stream_handler.setFormatter(log_handler_format)
        logger.addHandler(log_stream_handler)
    return logger

# コマンド呼び出し
if __name__ == '__main__':
    # 初期化
    rc = 0
    logger = setup_logger(__name__, logging.DEBUG)  # logging.INFO
    # 実行
    try:
        # パラメタチェック
        import argparse
        p = argparse.ArgumentParser()
        p.add_argument('-l', '--site_list', action='store_true', help='定義されているサイト（自治体名）を出力する。')
        p.add_argument('names', nargs='*', type=str, help='自治体名を指定する。省略した場合は全自治体を対象とする。')
        args = p.parse_args(sys.argv[1:])
        names = args.names
        # 開始
        msg = 'crawler.py start.'
        logger.info(msg)
        print(msg, file=sys.stderr)
        # 取得開始
        cobj = Crawler()
        names_all = cobj.get_names()
        if args.site_list:
            print('定義済自治体名：' + str(names_all), file=sys.stderr)
            cobj = None
            sys.exit(0)

        if len(names) == 0:
            names = names_all
        for name in names:
            if name not in names_all:
                raise Exception('指定された自治体名が存在しません。' + str(name))
            print('【' + name + '】', file=sys.stderr)
            rc = cobj.download_datasets_in_site(name)
            if rc != 0:
                break
        cobj = None
    except Exception as e:
        logger.exception(e)
        print('', file=sys.stderr)
        rc = 99
    # 終了
    msg = 'crawler.py ended, rc=' + str(rc)
    logger.info(msg)
    print(msg, file=sys.stderr)

    # 復帰
    sys.exit(rc)

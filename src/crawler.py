#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# crawler.py: オープンデータを取得して共通形式でキャッシュする
# Copyright (C) N.Togashi 2025

import os
import sys
import datetime
import time
import re
import csv
import json
import copy
import pathlib
import requests
from urllib3 import ssl
import unicodedata
import openpyxl
import xlrd
from bs4 import BeautifulSoup
import logging

MY_CONFIG = {
    'settings_file': 'crawler.json',
    'download_dir': 'download',
    'cache_dir': 'cache',
    'dir_name': '{code}_{name}'
}
CONTENT_HEADER_XLS  = bytes.fromhex('d0cf11e0a1b11ae1')
CONTENT_HEADER_XLSX = bytes.fromhex('504b03041400')
HEADER_ROWS = 3
LINKDATA_URL_BASE = 'http://linkdata.org'
LINKDATA_DOWNLOAD_CONTENT = """<!DOCTYPE html>
<!--

    _/  _/            _/              _/              _/
   _/      _/_/_/    _/  _/      _/_/_/    _/_/_/  _/_/_/_/    _/_/_/        _/_/    _/  _/_/    _/_/_/
  _/  _/  _/    _/  _/_/      _/    _/  _/    _/    _/      _/    _/      _/    _/  _/_/      _/    _/
 _/  _/  _/    _/  _/  _/    _/    _/  _/    _/    _/      _/    _/      _/    _/  _/        _/    _/
_/  _/  _/    _/  _/    _/    _/_/_/    _/_/_/      _/_/    _/_/_/  _/    _/_/    _/          _/_/_/
                                                                                                 _/
                                                                                            _/_/
Link and Publish your data as RDF to the Linked Open Data Community
-->"""
XHTML_PAGE = '<!DOCTYPE html>'
FLOAT_FORMAT_RE = re.compile('([0-9]+\\.[0-9]+)')

class Crawler:

    APP_DIR = None
    state = None        # 都道府県名
    name = None         # 市区町村名
    code = None         # 市区町村コード
    names = []
    download_dir = None
    cache_dir = None
    locality_dict = {}
    __data_name_dict = {}
    __requests_retry_max = 3
    __requests_retry_seconds = 5
    __request_headers = {}
    __proxies = {}
    __geocode_url = 'https://maps.googleapis.com/maps/api/geocode/json' \
                    '?address={address}&language=ja&components=country:JP&key={key}'
    __google_maps_api_key = 'AIzaSyBZa9fI3N-L1OHnkiaGQODmOcPRP-HaWlA'
    __geocode_cache = None
    __geocode_cache_file = '.geocode_cache_{}.json'
    __geocode_cache_path = None
    __geocode_cache_update = False

    def __init__(self):
        """
        コンストラクタ
        :param index_path: index.jsonファイルのパス
        :return: なし
        :raise: Exception
        """
        logger = logging.getLogger(__name__)
        logger.info('__init__() start.')
        self.APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.BASE_DIR = self.APP_DIR
        if 'BASE_DIR' in os.environ:
            self.BASE_DIR = os.environ['BASE_DIR']
        self.settings_path = os.path.join(self.APP_DIR, 'src', MY_CONFIG['settings_file'])
        if not os.path.exists(self.settings_path):
            raise Exception('settings path not found, file='+self.settings_path)
        with open(self.settings_path, 'r') as fh:
            self.settings_json = json.loads(fh.read())
        self.LOCALITY_CODE_FILE = self.settings_json['common']['locality_code_file']
        self.IGNORE_WORD_IN_DATASET_NAME = \
            self.settings_json['common']['ignore_word_in_dataset_name']
        self.KIND_LIST_NORMALIZED_RE = \
            {k:re.compile(v) for k, v in \
                self.settings_json['common']['kind_list_normalized_re'].items()}
        self.OPENDATA_SITES = self.get_opendata_sites()
        self.site_names = list(self.OPENDATA_SITES.keys())
        logger.debug('site_names=' + str(self.site_names))
       	self.id_item_list = self.settings_json['common']['id_item_list']
        self.name_item_list_1 = self.settings_json['common']['name_item_list_1']
        self.name_item_list_2 = self.settings_json['common']['name_item_list_2']
        self.address_item_list = self.settings_json['common']['address_item_list']
        self.lat_item_list = self.settings_json['common']['lat_item_list']
        self.lng_item_list = self.settings_json['common']['lng_item_list']
        # download_dirの確認
        self.download_dir = os.path.join(self.BASE_DIR, MY_CONFIG['download_dir'])
        if not os.path.exists(self.download_dir):
            os.makedirs(self.download_dir)
            logger.info('makedirs download_dir ' + self.download_dir)
        # cache_dirの確認
        self.cache_dir = os.path.join(self.BASE_DIR, MY_CONFIG['cache_dir'])
        if not os.path.exists(self.cache_dir):
            os.makedirs(self.cache_dir)
            logger.info('makedirs cache_dir ' + self.cache_dir)
        # ファイルパスの設定
        self.LOCALITY_CODE_PATH = os.path.join(self.APP_DIR,
                MY_CONFIG['download_dir'], self.LOCALITY_CODE_FILE)
        logger.info('__init__() ended.')
        # 復帰
        return

    def __del__(self):
        """
        デストラクタ
        :return: なし
        """
        logger = logging.getLogger(__name__)
        logger.info('__del__() start.')
        pass
        logger.info('__del__() ended.')
        return

    def crawler_main(self, names):
        """
        クローラーのメイン処理
        :param names: list型、サイト名文字列のリスト
        :return: int型、復帰値、0=正常終了、他=異常終了
        :raise: Exception型、メッセージ文字列
        """
        logger = logging.getLogger(__name__)
        logger.info('crewler_main() start, names=' + str(names))
        # 実行
        rc = 0
        if names is None or not isinstance(names, list) \
        or not all([v.split('/')[0] in self.site_names for v in names]):
            msg = 'namesパラメタ値に誤りがあります。'
            logger.error(msg+'names='+str(names))
            raise Exception(msg)

        for nno, name in enumerate(names):
            print(str(nno+1)+'/'+str(len(names))+' 【'+name+'】', \
                    file=sys.stderr, end='')
            logger.debug('nno='+str(nno)+', name='+str(name))
            name_pkgs = name.split('/')
            name = name_pkgs[0]
            site_info = self.get_site_info(name)
            print('サイト=' + site_info['code'], site_info['name'] \
                    + '：' + site_info[site_info['webapi']]['api_url'], file=sys.stderr)
            self.__request_headers = {}
            if 'http_request_headers' in site_info[site_info['webapi']]:
                for k, v in site_info[site_info['webapi']]['http_request_headers'].items():
                    self.__request_headers[k] = v
                logger.info('self.__request_headers: ' + str(self.__request_headers))
            rc = self.cache_initialize(site_info)
            if len(name_pkgs) > 1:
                packageids = name_pkgs[1:]
            else:
                packageids = self.get_packageids_in_site(site_info)
            # マップ情報一覧を初期化する
            map_list_name = {}
            map_list_path = os.path.join(self.packages_dir, '.map_list')
            for pno, packageid in enumerate(packageids):
                if not isinstance(packageid, str):
                    packageid = str(packageid)
                print(str(pno+1)+'/'+str(len(packageids))+' ', \
                        file=sys.stderr, end='')
                logger.debug('pno='+str(pno)+', packageid='+str(packageid))
                map_list = []
                package_info = self.get_package_info(site_info, packageid)
                if package_info is None:
                    msg = 'パッケージの取得に失敗しました。id=' + packageid
                    print(msg, file=sys.stderr)
                    logger.debug(msg)
                    continue
                msg = self.select_package_info(site_info, packageid, package_info)
                if msg is not None:
                    print(msg, file=sys.stderr)
                    continue
                package_title = self.get_package_title(package_info)
                print(package_title + '：', file=sys.stderr, end='')
                package_kind = self.get_package_kind(package_info)
                if package_kind is None:
                    msg = '対象外のパッケージです。'
                    print(msg, file=sys.stderr)
                    logger.debug(msg)
                    continue
                resources = self.get_all_resources(package_info)
                resources = self.get_valid_resources(resources)
                if len(resources) == 0:
                    msg = '対象のリソースが存在しません。'
                    print(msg, file=sys.stderr)
                    logger.debug(msg)
                    continue
                self.package_dir = os.path.join(self.packages_dir, packageid)
                if not os.path.exists(self.package_dir):
                    os.mkdir(self.package_dir)
                    logger.info('mkdir locality directory=' + self.package_dir)
                for rno, resource in enumerate(resources):
                    logger.debug('rno='+str(rno)+', name='+str(resource['name']))
                    resource = self.get_content_by_resource(package_kind, resource)
                    if resource is None or resource['_content_'] is None:
                        print('リソースがダウンロードできません。', file=sys.stderr)
                        continue
                    # HTML形式でダウンロードページならダウンロードする
                    resource = self.check_html_page(resource, self.package_dir)
                    if resource is None or resource['_content_'] is None:
                        print('HTMLページからダウンロードできません。', file=sys.stderr)
                        continue
                    # コンテンツの内容でformatを検査する
                    resource = self.check_content_format(resource, self.package_dir)
                    if resource is None or resource['_content_'] is None:
                        print('扱えない形式のコンテンツです。', file=sys.stderr)
                        continue
                    print(resource['format']+':', end='', file=sys.stderr)
                    # コンテンツを表形式に変換する
                    resource = self.content_to_table(resource, self.package_dir)
                    if resource is None or resource['_table_'] is None:
                        print('表データに変換できません。', file=sys.stderr)
                        continue
                    # 表形式の項目名からマップ情報を作成する
                    resource = self.make_map_from_table(resource, package_kind)
                    if resource is None or resource['_map_'] == {}:
                        print('マップ情報が作成できません,' \
                                + resource['_map_msg_'], file=sys.stderr)
                        continue
                    if resource['_map_'] != None:
                        map_list.append(resource['_map_'])
                    # 表形式をマップデータ(JSON形式)に変換する
                    map_data = self.table_to_mapdata(resource)
                    # マップデータをキャッシュに保存する
                    self.save_to_cache(map_data, package_title+'_'+str(rno), \
                            dir=packageid)
                    print('['+package_kind+']', 'マップデータを作成しました。', file=sys.stderr)

                map_list_name[packageid] = map_list

            # マップ情報をサイト単位に保存する
            with open(map_list_path, 'w') as hfm:
                hfm.write(json.dumps(map_list_name, \
                        ensure_ascii=False).replace('[], "', '[],\n"').replace( \
                        '}], "', '}],\n"'))
        # 復帰
        logger.info('crewler_main() ended, rc=' + str(rc))
        return rc

    def get_opendata_sites(self):
        """
        crawler.jsonの有効なサイト情報を取得する。
        """
        logger = logging.getLogger(__name__)
        logger.info('get_opendata_sites() start.')
        # 実行
        sites = {}
        for k, v in self.settings_json['local_gov'].items():
            if 'skip' in v and v['skip']:
                continue
            sites[k] = v
        # 復帰
        logger.info('get_opendata_sites() ended, sites=' + str(sites))
        return sites

    def get_names(self):
        """
        サイト名一覧を取得する。
        :return: リスト型、文字列のサイト名リスト
        """
        logger = logging.getLogger(__name__)
        logger.info('get_names() start.')
        # 復帰
        logger.info('get_names() ended, rc=' + str(self.site_names))
        return self.site_names

    def url_get(self, url, file, cache=True, dir=None):
        """
        指定されたURLの内容を取得してdownload_dirに保存する
        :param url: str型、取得する資源のURL
        :param file: str型、URLのファイル名部分
        :param cache: boolean型、True=キャッシュを使用する、False:キャッシュを使用しない
        :return: str型、取得したコンテンツ、取得できない場合はNoneを返却する
        :raise: Exception
        """
        logger = logging.getLogger(__name__)
        logger.info('url_get() start, url=' + str(url) + ', file=' + str(file))
        # 実行
        content = None
        if url is None or not isinstance(url, str) or url == '':
            msg = 'urlパタメタは無効な値です。'
            logger.error(msg+str(url))
            # raise Exception(msg)
            logger.debug('url_get() ended_1.')
            return content
        if file is None or not isinstance(file, str) or file == '':
            msg = 'fileパタメタは無効な値です。'
            logger.error(msg+str(file))
            # raise Exception(msg)
            logger.debug('url_get() ended_2.')
            return content
        content = None
        cache_dir = dir
        if cache_dir == None:
            cache_dir = self.download_dir
        if cache \
        and os.path.exists(os.path.join(cache_dir, file)):
            # キャッシュから読み込む
            with open(os.path.join(cache_dir, file), 'rb') as hf:
                content = hf.read()
            logger.debug('Hit cache, file=' + os.path.join(cache_dir, file))
        else:
            if False:
                # テスト用に迂回する場合
                msg = 'no Hit in cache, file=' + os.path.join(cache_dir, file)
                logger.debug(msg)
                print(msg+', url='+url, file=sys.stdout)
                content = msg
                logger.debug('url_get() ended_3.')
                return content
            # HTTP GET
            time.sleep(1)
            for try_i in range(self.__requests_retry_max):
                try_ok = False
                try:
                    ssl_context = ssl.create_default_context()
                    ssl_context.set_ciphers("DEFAULT:!aNULL:!eNULL:!MD5:!3DES:!DES:!RC4:!IDEA:!SEED:!aDSS:!SRP:!PSK")
                    session = requests.session()
                    adapter = requests.adapters.HTTPAdapter()
                    adapter.init_poolmanager(1, 1,  ssl_context=ssl_context)
                    session.adapters.pop("https://", None)
                    session.mount("https://", adapter)
                    res = session.get(url, headers=self.__request_headers, \
                            proxies=self.__proxies, timeout=(15.0, 60.0))
                    try_ok = True
                except TimeoutError as e:
                    logger.exception(e)
                    res = None
                    break
                except requests.exceptions.ProxyError as e:
                    logger.exception(e)
                    res = None
                    if (try_i+1) < self.__requests_retry_max:
                        time.sleep(self.__requests_retry_seconds)
                except Exception as e:
                    logger.exception(e)
                    res = None
                    if (try_i+1) < self.__requests_retry_max:
                        time.sleep(self.__requests_retry_seconds)
                if try_ok:
                    break
            if res is None or res.status_code != requests.codes.ok:
                logger.error('requests error(2): ' + str(res))
            else:
                logger.debug('response.encoding: ' + str(res.encoding))
                content = res.content
            if res is not None:
                res.close()
            # キャッシュディレクイトリに保存する
            if cache and content is not None:
                with open(os.path.join(cache_dir, file), 'wb') as f:
                    f.write(content)
            # logger.debug('OK requests.get(), size=' + str(len(content)))

        # 復帰
        logger.info('url_get() ended, file=' + str(file))
        return content

    def convert_to_utf8(self, content, rows=0):
        """
        byte配列型文字列をUTF-8に変換する
        :param content: byte配列
        :param rows: int型、変換する行数
        :return: str型、UTF-8文字列
        """
        logger = logging.getLogger(__name__)
        logger.info('convert_to_utf8() start, rows=' + str(rows))
        # 実行
        data = None
        cont = content
        if rows > 0:
            lines = []
            index = 0
            while len(lines) < (rows+15):
                try:
                    n_index = content[index:].index(b'\n')
                    lines.append(content[index:index+n_index+1])
                    index += n_index + 1
                except Exception as e:
                    lines.append(content[index:])
                    break
            cont = lines[0]
            for i in range(1, len(lines)):
                cont += lines[i]
        res = {'encoding': 'utf-8', 'confidence': 0.5, 'language': 'Japanese'}
        try:
            res = detect(cont)
        except Exception as e:
            logger.exception(e)
        logger.debug('detect=' + str(res))
        if res['encoding'] is None \
        or not (res['encoding'].upper() == 'UTF-8' or res['encoding'].upper() == 'UTF-16'):
            res['encoding'] = 'CP932'
        try:
            data = content.decode(res['encoding'], errors='replace')
        except Exception as e:
            logger.exception(e)
            data = str(content)
        # 復帰
        logger.info('convert_to_utf8() ended, rc=' + str(len(data)))
        return data

    def data_from_csv(self, content, info):
        """
        CSV形式のデータをJSON形式に変換する。
        :param content: str型、CSV形式文字列
        :param info: dict型、形式情報
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.info('data_from_csv() start.')
        # 実行
        table = self.table_from_csv(content)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.info('data_from_csv() ended.')
        return map_data

    def data_from_tsv(self, content, info):
        """
        TSV形式のデータをJSON形式に変換する。
        :param content: str型、TSV形式文字列
        :param info: dict型、形式情報
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.info('data_from_tsv() start.')
        # 実行
        table = self.table_from_tsv(content)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.info('data_from_tsv() ended.')
        return map_data

    def data_from_xls(self, content, info, file):
        """
        Excel形式(XLS)のデータをJSON形式に変換する。
        :param content: byte配列、Excel形式データ
        :param info: dict型、形式情報
        :param file: str型、Excel形式データのファイルパス
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.info('data_from_xls() start.')
        # 実行
        table = self.table_from_xls(file)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.info('data_from_xls() ended.')
        return map_data

    def data_from_xlsx(self, content, info, file):
        """
        Excel形式(XLSX)のデータをJSON形式に変換する。
        :param content: byte配列、Excel形式データ
        :param info: dict型、形式情報
        :param file: str型、Excel形式データのファイルパス
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.info('data_from_xlsx() start.')
        # 実行
        table = self.table_from_xlsx(file)
        map_data = self.table_to_mapdata(table, info)
        # 復帰
        logger.info('data_from_xlsx() ended.')
        return map_data

    def table_to_mapdata(self, resource):
        """
        二次元配列データ（行、列）をJSON形式データに変換する。
        :param resource: dict型、マップ情報と二次元配列データを含むリソース情報
        :return: str型、JSON形式データ
        """
        logger = logging.getLogger(__name__)
        logger.info('table_to_mapdata() start.')
        # 実行
        if resource is None or not isinstance(resource, dict) \
        or '_table_' not in resource or not isinstance(resource['_table_'], list) \
        or '_map_' not in resource or not isinstance(resource['_map_'], dict):
            msg = 'resourceパラメタは無効な値です。'
            logger.error(msg+str(resource))
            raise Exception(msg)
        map_data = []
        data = resource['_table_']
        info = resource['_map_']
        # 名称リスト初期化
        self.__data_name_dict = {}
        for i, rec in enumerate(data):
            if len(rec) == 0 \
            or (type(rec[0]) == str and (len(rec[0]) == 0 or rec[0][0] == '#')):
                continue
            if i >= info['header'] and len(rec) > 0 \
            and (type(rec[0]) != str or (len(rec[0]) > 0 and rec[0][0] != '#')):
                name = ''
                for n in info['name']:
                    if type(n) == int:
                        if len(rec) > n:
                            name += str(rec[n])
                    else:
                        name += str(n)
                if name == '' or name == '○' or name == '◎':
                    continue
                if name in self.__data_name_dict:
                    self.__data_name_dict[name] += 1
                else:
                    self.__data_name_dict[name] = 1
        # テーブルデータをマップ情報に変換する
        headers = list(data[info['header']-1])
        for i in range(len(headers)):
            if headers[i] == '':
                headers[i] = '#' + ('000'+str(i+1))[-4:]
        no = 0;
        for i in range(len(data)):
            if i < info['header']:
                continue
            if len(data[i]) == 0 or (type(data[i][0]) == str \
                    and (len(data[i][0]) == 0 or data[i][0][0] == '#')):
                continue
            if len(data[i]) > 0 and isinstance(data[i][0], str) \
            and data[i][0][:3] == '記入例':
                continue
            name = ''
            for n in info['name']:
                if type(n) == int:
                    if len(data[i]) > n:
                        name += str(data[i][n])
                else:
                    name += str(n)
            if name == '' or name == '○' or name == '◎':
                continue
            info_items = []
            for j in info['info']:
                if len(data[i]) <= j:
                    continue
                if data[i][j] == '':
                    continue
                info_items.append({headers[j]: data[i][j]})
            # info_items.append({'データセットURL': info['url']})
            error = ''
            no += 1
            id_value = ('000' + str(no))[-4:]
            if info['id'] >= 0:
                if data[i][info['id']] != '':
                    id_value = data[i][info['id']]
            else:
                error += 'id未設定。'
            loc_name = self.state + self.name
            if self.state == self.name:
                loc_name = self.name
            address = ''
            if info['address'] >= 0 and len(data[i]) > info['address']:
                address = data[i][info['address']]
            (lat, lng, msg) = self.lat_lng_from_data(data[i], info, address)
            data_value = {
                    "id": id_value,
                    "locality_code": self.code,
                    "locality_name": loc_name,
                    "kind": info['kind'],
                    "dataset": resource['_package_']['_info_']['title'],
                    "label": name,
                    "lat": lat,
                    "lng": lng,
                    "info": info_items,
            }
            if msg != '':
                data_value["error"] = msg
            map_data.append(data_value)
        self.__data_name_dict = {}
        # 復帰
        logger.info('table_to_mapdata() ended, rc=' + str(len(map_data)))
        return map_data

    def exists_in_cache(self, file, dir=None):
        """
        指定されたファイルがキャッシュに存在するか確認する
        :param file: str型、ファイル名
        :return: boolean型、True=存在する、False=存在しない
        """
        logger = logging.getLogger(__name__)
        logger.info('exists_in_cache() start, file=' + str(file))
        # 実行
        rc = False
        cache_dir = dir
        if cache_dir is None:
            cache_dir = self.cache_dir
        if os.path.exists(os.path.join(cache_dir, file)):
            rc = True
        # 復帰
        logger.info('exists_in_cache() ended, rc=' + str(rc))
        return rc

    def save_to_cache(self, mapping_data, file, dir=None):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('save_to_cache() start, file=' + str(file))
        # 実行
        content = None
        if mapping_data is None or not isinstance(mapping_data, list):
            msg = 'mapping_dataパラメタは無効な値です。'
            logger.error(msg+str(mapping_data))
            raise Exception(msg)
        if file is None or not isinstance(file, str) \
        or file == '':
            msg = 'fileパラメタは無効な値です。' 
            logger.error(msg+str(file))
            raise Exception(msg)
        if dir is not None and (not isinstance(dir, str) or dir == ''):
            msg = 'dirパラメタは無効な値です。'
            logger.error(msg+str(dir))
            raise Exception(msg)
        cache_dir = self.cache_dir
        if hasattr(self, 'cache_package_dir') and self.cache_package_dir != None:
            cache_dir = self.cache_package_dir
        if dir is not None:
            if dir[0] == '/' or dir[0] == '.':
                cache_dir = dir
            else:
                cache_dir = os.path.join(cache_dir, dir)
        save_path = os.path.join(cache_dir, file)
        try:
            if not os.path.exists(cache_dir):
                os.makedirs(cache_dir)
            content = json.dumps(mapping_data, ensure_ascii=False).replace('}, {', '},\n{')
            with open(save_path, 'w') as f:
                f.write(content)
        except Exception as e:
            logger.exception(e)
            logger.error('path=' + str(save_path))
            logger.error('mapping_data=' + str(mapping_data))
            raise Exception('error in save_to_cache(), path=' + save_path)
        # 復帰
        logger.info('save_to_cache() ended, path=' + str(save_path))
        return save_path

    def __save_to_download_dir(self, content, file):
        """
        contentをdownload_dirにファイルとして保存する
        :param content: byte型、ファイルの内容
        :param file: str型、ファイル名
        :return: boolean型、True=正常終了
        :raise Exception: 異常終了
        """
        logger = logging.getLogger(__name__)
        logger.info('__save_to_download_dir() start, file=' + file)
        logger.debug('save file=' + os.path.join(self.download_dir, file))
        # 実行
        try:
            with open(os.path.join(self.download_dir, file), 'wb') as f:
                f.write(content)
        except Exception as e:
            logger.exception(e)
            raise Exception('error in store_to_download_dir().')
        # 復帰
        logger.info('__save_to_download_dir() ended.')
        return True

    def lat_lng_from_data(self, data, info, address):
        """
        レコードの緯度・経度を浮動小数点形式で返却する。
        緯度・経度が設定されていない場合は、名称または住所から緯度・経度を求める。
        :param data: Array型、データレコード
        :param info: dict型、データ定義
        :param address: str型、住所
        :return: (緯度, 経度, メッセージ)、緯度=float型、経度=float型、メッセージ=str型
        """
        logger = logging.getLogger(__name__)
        # logger.debug('lat_lng_from_data() start.')
        # logger.debug('lat_lng_from_data() start, data='+str(data) \
        #         +', info='+str(info)+', address='+str(address))
        # logger.debug('self.state='+self.state+', self.name='+self.name)
        # 実行
        lat = ''
        lng = ''
        msg = ''
        name = ''
        for n in info['name']:
            if type(n) == int:
                name += str(data[n])
            else:
                name += str(n)
        if name in self.__data_name_dict \
        and self.__data_name_dict[name] > 1:
            msg += '名称が重複。'
            logger.debug('名称が重複, name=' + name + ', count=' \
                    + str(self.__data_name_dict[name]))
        if info['lat'] == -1 or info['lng'] == -1:
            msg += '緯度・経度が未定義。'
        elif len(data) <= info['lat'] or len(data) <= info['lng']:
            msg += '緯度・経度の値無し。'
        elif data[info['lat']] == '' or data[info['lng']] == '':
            msg += '緯度・経度が未設定。'
        else:
            lat = data[info['lat']]
            lng = data[info['lng']]

        if lat == '' or lng == '':
            # 住所から緯度・経度を取得する
            if address == '':
                msg += '住所未設定。'
            else:
                # logger.debug('address='+str(address))
                param_addr = address
                if address[:len(self.state)] == self.state:
                    if address[len(self.state):len(self.state)+len(self.name)] == self.name:
                        # 住所に都道府県名と市区町村名が含まれている→加工不要
                        pass
                    else:
                        # 住所に都道府県名が含まれているが市区町村名は含まれていない
                        # → 市区町村名を挿入する
                        if self.state == self.name:
                            # 都道ぬ件名が重複指定された場合→加工不要
                            pass
                        else:
                            param_addr = self.state + self.name + address[len(self.state):]
                else:
                    if address[:len(self.name)] == self.name:
                        # 住所に市区町村名が含まれている→都道府県名を追加する
                        param_addr = self.state + address
                    else:
                        # 住所に都道府県名も市区町村名も含まれていない
                        # → 都道府県名と市区町村名を追加する
                        if self.state == self.name:
                            # 都道ぬ件名が重複指定された場合→都道府県名のみを追加する
                            param_addr = self.state + address
                        else:
                            param_addr = self.state + self.name + address
                # logger.debug('param_addr='+str(param_addr))
                if param_addr in self.__geocode_cache:
                    result = self.__geocode_cache[param_addr]
                    logger.debug('Hit in __geocode_cache, 住所=' + param_addr)
                else:
                    if True:	# ToDo:
                        result = {'status': 'NG', \
                                'note': '住所から緯度・経度への変換は中止' \
                                '（Google Mapsの課金が発生するため）'}
                    else:
                        geo_url = self.__geocode_url.format(address=param_addr, \
                                key=self.__google_maps_api_key)
                        content = self.url_get(geo_url, None, cache=False)
                        if content is None:
                            result = {'status': 'NG'}
                        else:
                            result = json.loads(content)
                            self.__geocode_cache[param_addr] = result
                            self.__geocode_cache_update = True
                # logger.debug(param_addr + '：' + str(result))
                if result['status'] == 'OK':
                    if len(result['results']) == 1:
                        res = result['results'][0]
                        if 'location_type' in res['geometry']:
                            # if res['geometry']['location_type'] \
                            #       in ['ROOFTOP', 'RANGE_INTERPOLATED', \
                            #         'GEOMETRIC_CENTER', 'APPROXIMATE']:
                            if self.state + self.name in res['formatted_address']:
                                lat = res['geometry']['location']['lat']
                                lng = res['geometry']['location']['lng']
                                if res['geometry']['location_type'] == 'ROOFTOP':
                                    msg += '住所から緯度・経度を取得(' \
                                            + res['geometry']['location_type'] + ')。'
                                else:
                                    msg += '住所から近隣の緯度・経度を取得(' \
                                            + res['geometry']['location_type']
                                    msg += ')=' + res['formatted_address'] + '。'
                        else:
                            msg += '住所：' + param_addr + '：NG geometry無し。'
                    else:
                        msg += '住所：' + param_addr + '：NG 複数results。'
                else:
                    msg += '住所：' + param_addr + '：NG status=' + result['status'] + '。'
        if (lat == '' or lng == '') and address != '':
            msg += '住所から取得できず。'

        if lat == '' or lng == '':
            # 名称から緯度・経度を取得する
            if self.state == self.name:
                param_name = self.state + ' ' + name
            else:
                param_name = self.state + self.name + ' ' + name
            # logger.debug('param_name=' + param_name)
            if param_name in self.__geocode_cache:
                result = self.__geocode_cache[param_name]
                logger.debug('Hit in __geocode_cache, 名称=' + param_name)
            else:
                if True:	# ToDo:
                    result = {'status': 'NG', \
                            'note': '名称から緯度・経度への変換は中止' \
                            '（Google Mapsの課金が発生するため）'}
                else:
                    geo_url = self.__geocode_url.format(address=param_name, \
                            key=self.__google_maps_api_key)
                    content = self.url_get(geo_url, None, cache=False)
                    if content is None:
                        result = {'status': 'NG'}
                    else:
                        result = json.loads(content)
                        self.__geocode_cache[param_name] = result
                        self.__geocode_cache_update = True
            # logger.debug(param_name + '：' + str(result))
            if result['status'] == 'OK':
                for res in result['results']:
                    if 'geometry' in res:
                        if 'location_type' in res['geometry']:
                            if self.state + self.name not in res['formatted_address']:
                                continue
                            if res['geometry']['location_type'] == 'ROOFTOP':
                                lat = res['geometry']['location']['lat']
                                lng = res['geometry']['location']['lng']
                                msg += '名称から緯度・経度を取得1(' \
                                        + res['geometry']['location_type']
                                if unicodedata.normalize('NFKC', \
                                        res['address_components'][0]['long_name']) \
                                        == unicodedata.normalize('NFKC', name):
                                    msg += ')。'
                                else:
                                    msg += ')=' + res['formatted_address'] + '。'
                                    break
                            elif res['geometry']['location_type'] == 'GEOMETRIC_CENTER':
                                if res['address_components'][0]['long_name'] == name:
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称から緯度・経度を取得2(' \
                                            + res['geometry']['location_type'] + ')。'
                                    break
                                elif len(result['results']) == 1 \
                                and (res['address_components'][0]['long_name'][:2] \
                                        == name[:2] \
                                  or res['address_components'][0]['long_name'][-2:] \
                                        == name[-2:]):
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称（類似）から緯度・経度を取得3(' \
                                            + res['geometry']['location_type']
                                    if unicodedata.normalize('NFKC', \
                                            res['address_components'][0]['long_name']) \
                                            == unicodedata.normalize('NFKC', name):
                                        msg += ')。'
                                    else:
                                        msg += ')=' + res['formatted_address'] + '。'
                                    # ToDo: msg += '住所未設定。'
                                    break
                            elif address != '':
                                a_addr = unicodedata.normalize('NFKC', \
                                        address.replace('番地', '−'))
                                # logger.debug('a_addr=' + a_addr)
                                a_match = re.match( \
                                        '[^0-9０-９]*([0-9０-９\\-−－ー]*)$', a_addr)
                                if a_match is not None and a_match.group(1) != '':
                                    # logger.debug(str(a_match.groups()))
                                    r_addr = unicodedata.normalize('NFKC', \
                                            res['formatted_address'].replace('番地', '−'))
                                    # logger.debug('r_addr=' + r_addr)
                                    r_match = re.match( \
                                            '[^0-9０-９]*([0-9０-９\\-−－ー]*)$', r_addr)
                                    if r_match is not None and r_match.group(1) != '':
                                        if a_match.group(1) == r_match.group(1):
                                            lat = res['geometry']['location']['lat']
                                            lng = res['geometry']['location']['lng']
                                            msg += '名称から緯度・経度を取得4(' \
                                                    + res['geometry']['location_type'] + ')'
                                            msg += '=番地が同一(' + a_match.group(1) + ')。'
                                            break
                            elif address == '' and len(result['results']) == 1:
                                if res['formatted_address'].find(self.state) >= 0 \
                                and res['formatted_address'].find(self.name) >= 0:
                                    lat = res['geometry']['location']['lat']
                                    lng = res['geometry']['location']['lng']
                                    msg += '名称から近隣の緯度・経度を取得5(' \
                                            + res['geometry']['location_type'] \
                                         + ')=' + res['formatted_address'] + '。'
                                    # ToDo: msg += '住所未設定。'
                                    break
                            else:
                                msg += '名称：' + param_name \
                                        + '：NG location_type=' \
                                        + res['geometry']['location_type'] + '。'
                        else:
                            msg += '名称：' + param_name + '：NG location_type無し。'
                    else:
                        msg += '名称：' + param_name + '：NG geometry無し。'
            else:
                msg += '名称：' + param_name + '：NG status=' + result['status'] + '。'

        # 浮動小数点形式に変換する
        if type(lat) != float:
            (lat, m) = self.string_to_float(lat)
            if m != '':
                msg += '緯度：' + m
        if type(lng) != float:
            (lng, m) = self.string_to_float(lng)
            if m != '':
                msg += '経度：' + m
        # 復帰
        # logger.debug('lat_lng_from_data() ended, lat=' + str(lat) \
        #         + ', lng=' + str(lng) + ', msg=' + msg)
        return (lat, lng, msg)

    def string_to_float(self, v_str):
        """
        文字列形式の緯度・経度を浮動小数点に変換する
        :param v_str: str型、文字列形式の緯度または経度
        :return: 浮動小数点形式の緯度または経度。ただし、文字列形式の値が空文字の場合は空文字を返却する
        """
        msg = ''
        msg_correct = ''
        v_float = ''
        try:
            if v_str is None:
                raise Exception('msg:値なし。')
            if v_str == '' or v_str == ' ' or v_str == '　' or v_str == '-' or v_str == '－':
                raise Exception('msg:値が空。')
            v_str = v_str.replace(' ', '').replace(',', '')
            match = FLOAT_FORMAT_RE.search(v_str)
            if match is not None:
                v_str = match.group(1)
            v_strs = v_str.split('.')
            if len(v_strs) == 2:
                # 正しい形式
                v_float = float(v_str)
            elif len(v_strs) == 1:
                # 「.」を入れ忘れた？
                if v_strs[0][0] == '1':
                    v_float = float(v_strs[0][:3] + '.' + v_strs[0][3:])
                else:
                    v_float = float(v_strs[0][:2] + '.' + v_strs[0][2:])
                msg += '度に小数点無し。'
            else:
                # 度.分.秒形式？、日本測地系から世界測地系に簡易補正
                v_float = float(v_strs[0]) \
                        + float(v_strs[1]) / 60.0
                msg += '度.分.秒形式から変換'
                if len(v_strs) >= 3:
                    if len(v_strs) == 3:
                        v_float += float(v_strs[2]) / 3600.0
                    else:
                        v_float += float(v_strs[2]+'.'+v_strs[3]) / 3600.0
                    # 沼津市地点での簡易補正（秒の指定が無ければ誤差の範囲とみなす）
                    if v_float > 90.0:
                        v_float -= 11.29 / 3600
                        msg_correct = '-11.29秒'
                    else:
                        v_float += 11.85 / 3600
                        msg_correct = '+11.85秒'
                    msg += '・簡易補正'
                msg += '(' + v_str + msg_correct + ')。'
        except Exception as e:
            if len(e.args) > 0 and type(e.args[0]) == str and e.args[0][:4] == 'msg:':
                msg += e.args[0][4:]
            else:
                logger = logging.getLogger(__name__)
                logger.exception(e)
                msg += str(e)
        return (v_float, msg)


    def get_site_info(self, name):
        """
        指定された名称の自治体定義を取得する。
        :param name: str型、自治体名
        :return: dict型、自治体定義
        :raise: Exception型、メッセージ
        """
        logger = logging.getLogger(__name__)
        logger.info('get_site_info() start.')
        # 実行
        site_info = None
        if name is None or name == '':
            msg = 'nameパラメタ（自治体名）を指定してください。'
            logger.error(msg+str(name))
            raise Exception(msg)
        if name not in self.site_names:
            msg = '指定された自治体名が存在しません。'
            logger.error(msg+str(name))
            raise Exception(msg)
        for key, value in self.OPENDATA_SITES.items():
            if key != name:
                # 対象外サイト
                continue
            site_info = value
            site_info['name'] = key
        if site_info is None:
            raise Exception('自治体定義が存在しません。' + str(name))
        self.name = name
        self.code = site_info['code']
        if self.name not in self.site_names:
            raise Exception('name not exists in index, ' + self.name)
        # self.state = self.name
        # self.code = [v['code'] for k,v in self.OPENDATA_SITES.items() if k==self.name][0]
        # logger.debug('state=' + self.state + ', code=' + self.code)
        self.packages_dir = os.path.join(self.download_dir, \
                    MY_CONFIG['dir_name'].format(code=self.code, name=self.name))
        # 復帰
        logger.info('get_site_info() ended, site_info=' + str(site_info))
        return site_info

    def get_packageids_in_site(self, site_info):
        """
        サイトのパッケージID一覧を取得する。
        :param names: dict型、サイト定義
        :return: dict型、パッケージIDリスト
        :raise: Exception型、メッセージ
        """
        logger = logging.getLogger(__name__)
        logger.info('get_packageids_in_site() start.')
        # 実行
        if site_info is None or not isinstance(site_info, dict) \
        or 'name' not in site_info:
            raise Exception('指定されたサイト定義が誤りです。')
        # パッケージリストを取得する
        self.packages_dir = os.path.join(self.download_dir, \
                MY_CONFIG['dir_name'].format(code=site_info['code'],
                name=site_info['name']))
        packages = self.get_package_list(site_info)

        # 復帰
        logger.info('get_packageids_in_site() ended, packages=' \
                + str(len(packages)))
        return packages

    def cache_initialize(self, site_info):
        """
        キャッシュを初期化する。
        :param site_info: dict型、サイト情報
        :return: int型、復帰コード、0=正常終了、他=異常終了
        """
        logger = logging.getLogger(__name__)
        logger.info('cache_initialize() start.')
        # 実行
        rc = 0
        self.state = site_info['name']
        self.code = site_info['code']
        logger.debug('state=' + self.state + ', code=' + self.code)
        self.packages_dir = os.path.join(self.download_dir, \
                MY_CONFIG['dir_name'].format(code=self.code, name=self.name))
        logger.debug('self.packages_dir=' + self.packages_dir)
        # geocode検索結果キャッシュを初期化
        self.__geocode_cache = {}
        self.__geocode_cache_update = False
        self.__geocode_cache_path = os.path.join(self.packages_dir, \
                self.__geocode_cache_file.format(site_info['name']))
        logger.debug('geocode_cache_path=' + self.__geocode_cache_path)
        if os.path.exists(self.__geocode_cache_path):
            with open(self.__geocode_cache_path, 'r') as hf:
                self.__geocode_cache = json.loads(hf.read())
                logger.info('use geocode cache, path=' + self.__geocode_cache_path)
        # キャッシュディレクトリを確認する
        self.cache_site_dir = os.path.join(self.cache_dir, \
                MY_CONFIG['dir_name'].format(code=site_info['code'], \
                        name=site_info['name']))
        if not os.path.exists(self.cache_site_dir):
            os.mkdir(self.cache_site_dir)
        # 対象地方自治体コード表を取得する
        self.locality_dict = {}
        with open(self.LOCALITY_CODE_PATH, 'r') as hf:
            hcsv = csv.reader(hf)
            self.locality_dict = \
                {rec[2]:rec[0] for rec in hcsv if rec[1] == site_info['name']}
        # 復帰
        logger.info('cache_initialize() ended, rc='+str(rc))
        return rc

    def get_locality_from_package(self, jpackage, default_code):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('get_locality_from_package() start')
        locality_name = ''
        locality_code = ''
        if ('groups' not in jpackage['result'] \
          or len(jpackage['result']['groups']) < 1) \
        and ('areas' not in jpackage['result'] \
          or len(jpackage['result']['areas']) < 1) \
        and ('organization' not in jpackage['result'] \
          or len(jpackage['result']['organization']) < 1):
            locality_name = self.state
        else:
            if 'groups' in jpackage['result'] \
            and len(jpackage['result']['groups']) > 0:
                i_group = 0
                if len(jpackage['result']['groups']) > 1:
                    if jpackage['result']['groups'][0]['id'] == 1:
                        i_group = 1
                if 'trailing_name' in jpackage['result']['groups'][i_group]:
                    locality_name = jpackage['result']['groups'][i_group]['trailing_name']
                    group_names = jpackage['result']['groups'][i_group]['name'].split('/')
                    if len(group_names) > 1 and locality_name == group_names[1]:
                        locality_name = group_name[0]
            if locality_name == '' \
            and 'areas' in jpackage['result']:
                locality_name = jpackage['result']['areas'][0]['name']
            if locality_name == '' \
            and 'organization' in jpackage['result']:
                if isinstance(jpackage['result']['organization']['name'], str) \
                and re.match('^[0-9a-zA-Z]+$', jpackage['result']['organization']['name']) == None:
                    locality_name = jpackage['result']['organization']['name']
                elif jpackage['result']['organization']['title'][-1] in ['市','区','町','村']:
                    if jpackage['result']['organization']['title'][:len(self.state)] == self.state:
                        locality_name = jpackage['result']['organization']['title'][len(self.state):]
                    else:
                        locality_name = jpackage['result']['organization']['title']
                    locality_name = locality_name.strip()
        if locality_name == '':
            locality_name = self.state
        if locality_name not in self.locality_dict:
            locality_code = default_code
        else:
            locality_code = self.locality_dict[locality_name]
        logger.info('get_locality_from_package() ended, locality_name=' \
                + str(locality_name) + ', locality_code=' + str(locality_code))
        return locality_name, locality_code

    def get_package_list(self, site):
        # パッケージリストを取得する
        logger = logging.getLogger(__name__)
        logger.info('get_package_list() start, site=' + str(site))
        # 実行
        content = None
        jpackages = None
        self.packages_dir = os.path.join(self.download_dir, \
                MY_CONFIG['dir_name'].format(code=site['code'], name=site['name']))
        if not os.path.exists(self.packages_dir):
            os.mkdir(self.packages_dir)
            logger.info('make site dir=' + self.packages_dir)
        packages_file = '.package_list.json'
        packages_path = os.path.join(self.packages_dir, packages_file)
        logger.debug('packages_path='+packages_path)
        if os.path.exists(packages_path):
            with open(packages_path, 'r') as fh:
                jpackages = json.loads(fh.read())
        else:
            if site['webapi'] == 'シラサギ' \
            and 'package_list_limit' in site[site['webapi']] \
            and site[site['webapi']]['package_list_limit'] == 0:
                url = site[site['webapi']]['api_url'] + 'package_list?limit=0&offset=0'
                content = self.url_get(url, packages_file, dir=self.packages_dir)
                if content is not None:
                    jpackages = json.loads(content)
            elif site['webapi'] == 'CKAN' \
            and 'package_list_limit' in site[site['webapi']] \
            and (site[site['webapi']]['package_list_limit'] == -1
              or site[site['webapi']]['package_list_limit'] == 0):
                url = site[site['webapi']]['api_url'] + 'package_list'
                if site[site['webapi']]['package_list_limit'] == 0:
                    url += '?limit=0&offset=0'
                content = self.url_get(url, packages_file, dir=self.packages_dir)
                if content is not None:
                    jpackages = json.loads(content)
            elif site['webapi'] == 'CKAN' \
            and  'package_list_path' in site[site['webapi']]:
                url = site[site['webapi']]['api_url'] \
                    + site[site['webapi']]['package_list_path']
                content = self.url_get(url, packages_file, dir=self.packages_dir)
                if content is not None:
                    jpackages = json.loads(content)
            else:
                offset = 0
                results = []
                while True:
                    url = site[site['webapi']]['api_url'] + 'package_list'
                    url += '?limit=' + str(site[site['webapi']]['package_list_limit']) \
                         + '&offset=' + str(offset)
                    content = self.url_get(url, packages_file, dir=self.packages_dir)
                    logger.info('content='+str(content)[:2048])
                    if content is None:
                        break
                    jpackages = json.loads(content)
                    if jpackages['success'] is not True:
                        break
                    results.extend(jpackages['result'])
                    offset += len(jpackages['result'])
                    if len(jpackages['result']) < site[site['webapi']]['package_list_limit']:
                        break
                    os.remove(os.path.join(self.packages_dir, packages_file))
                if len(results) > 0:
                    jpackages['result'] = results
                    content = json.dumps(jpackages, ensure_ascii=False)
                    with open(os.path.join(self.packages_dir, packages_file), 'w') as pfh:
                        pfh.write(content)
        # 復帰
        if jpackages is None:
            end_msg = 'None'
        else:
            end_msg = str(len(jpackages['result']))
        logger.info('get_package_list() ended, result=' + end_msg)
        return jpackages['result']

    def get_package_info(self, site_info, packageid):
        # パッケージ情報を取得する
        logger = logging.getLogger(__name__)
        logger.info('get_package_info() start, packageid=' + str(packageid))
        # 実行
        if site_info is None or not isinstance(site_info, dict) \
        or 'name' not in site_info:
            raise Exception('指定されたサイト定義が誤りです。')
        if packageid is None or not isinstance(packageid, str) \
        or packageid == '':
            raise Exception('指定されたパッケージIDが誤りです。')
        jpackage = None
        self.packages_dir = os.path.join(self.download_dir, \
                MY_CONFIG['dir_name'].format(code=site_info['code'], \
                name=site_info['name']))
        package_file = packageid+'.package'
        package_path = os.path.join(self.packages_dir, package_file)
        self.package_dir = os.path.join(self.packages_dir, packageid)
        self.cache_package_dir = os.path.join(self.cache_dir, \
                MY_CONFIG['dir_name'].format(code=site_info['code'], \
                name=site_info['name']))
        if os.path.exists(package_path):
            with open(package_path, 'r') as hf:
                jpackage = json.loads(hf.read())
        else:
            url = site_info[site_info['webapi']]['api_url'] + 'package_show?id=' + packageid
            content = self.url_get(url, package_file, dir=self.packages_dir)
            if content is not None:
                jpackage = json.loads(content)
                # logger.debug('jpackage=' + str(jpackage))
                if not 'success' in jpackage or not jpackage['success']:
                    logger.error('ERROR: error in result, result=' + str(jpackage))
                    # raise Exception('msg:パッケージ取得失敗：' + str(jpackage))
                    jpackage = None
                else:
                    if not 'result' in jpackage or not 'type' in jpackage['result'] \
                    or jpackage['result']['type'] != 'dataset':
                        logger.error('ERROR: error in dataset type, result=' \
                                + str(jpackage['result']))
                        # raise Exception('msg:datasetを含まないパッケージ：' + str(jpackage))
                        jpackage = None

        # 復帰
        logger.info('get_package_info() ended, jpackage=' + str(jpackage is not None))
        return jpackage

    def select_package_info(self, site_info, packageid, jpackage):
        """
        パッケージの内容を確認して選定する。
        :param site_info: dict型、サイト情報
        :param packageid: str型、パッケージID
        :param jpackage: dict型、JSON型パッケージ情報
        :return: str型、選定結果のメッセージ、None=正常、他=除外した理由
        """
        logger = logging.getLogger(__name__)
        logger.info('get_package_info() start, packageid=' + str(packageid))
        # 実行
        msg = None
        if packageid is None or not isinstance(packageid, str) \
        or packageid == '':
            msg = 'packageidパタメタは無効な値です。'
            logger.error(msg + str(packageid))
            raise Exception(msg)
        self.check_package_info(jpackage)

        # パッケージIDを取得する
        name = packageid
        if 'name' in jpackage['result']:
            name = jpackage['result']['name']
        # パッケージ名を取得する
        dataset_name = jpackage['result']['name']
        if 'title' in jpackage['result']:
            if len(jpackage['result']['title']) <= 128: 
                dataset_name = jpackage['result']['title']
        # 都道府県のパッケージの場合は市町村名を取得する
        locality_name, locality_code = self.get_locality_from_package( \
                jpackage, site_info['code'])
        self.name = locality_name
        self.code = locality_code
        # 正規化種別リストに従って種別を決定する
        kinds = [k for k,v in self.KIND_LIST_NORMALIZED_RE.items() \
                if v.search(dataset_name)]
        logger.debug('kinds=' + str(kinds))
        if len(kinds) == 0 \
        or (len(kinds) == 1 and kinds[0] == '##ignore##'):
            msg = dataset_name + '：種別が特定できないか無視対象です。' + str(kinds)
            # print(msg, file=sys.stderr)
            logger.debug(msg)
        else:
            if len(kinds) > 1 and kinds[0] == '##ignore##':
                kinds.pop(0)
            jpackage['_info_'] = {
                'name': name,
                'title': dataset_name,
                'kind': kinds[0]
            }
            name_ignore = [w for w in self.IGNORE_WORD_IN_DATASET_NAME \
                    if dataset_name.find(w) >= 0]
            if len(name_ignore) > 0:
                msg = dataset_name + '：特定単語を含むデータセットは除外。' \
                        + ','.join(name_ignore)
                # print(msg, file=sys.stderr)
                logger.debug(msg)

        # 復帰
        logger.info('select_package_info() ended, msg=' + str(msg))
        return msg

    def check_package_info(self, package_info):
        """
        パラメタで指定されたパッケージ情報のチェック
        :param package_info: dict型、パッケージ情報
        :return: bool型、True
        :raise: Exception型、エラーメッセージ
        """
        if package_info is None or not isinstance(package_info, dict) \
        or 'success' not in package_info:
            raise Exception('指定されたパッケージ情報が誤りです。')
        if not package_info['success'] or 'result' not in package_info \
        or not isinstance(package_info['result'], dict) \
        or 'resources' not in package_info['result'] \
        or not isinstance(package_info['result']['resources'], list):
            raise Exception('指定されたパッケージ情報は異常です。')
        return True

    def get_package_title(self, package):
        """
        パッケージのタイトルを取得する。
        本関数の呼び出し前に、select_package_info()関数を呼び出すこと。
        :param package: dict型、パッケージ情報
        :return: str型、タイトル
        """
        self.check_package_info(package)
        return package['_info_']['title']

    def get_all_resources(self, package):
        """
        パッケージ情報から全リソースの必要情報を取得する
        """
        logger = logging.getLogger(__name__)
        logger.info('get_all_resources() start.')
        # 実行
        self.check_package_info(package)
        resources = [{'name': v['name'] if 'name' in v else None,
                    'filename': v['filename'] if 'filename' in v else None,
                    'format': v['format'] if 'format' in v else None,
                    'mimetype': v['mimetype'] if 'mimetype' in v else None,
                    'created': v['created'] if 'created' in v else None,
                    'modified': v['modiried'] if 'modified' in v else None,
                    'url': v['url'] if 'url' in v else None,
                    'download_url': v['download_url'] if 'download_url' in v else None,
                    '_package_': package}
                for v in package['result']['resources']]
        # 不足情報を補う
        for resource in resources:
            if resource['url'] is None or resource['url'] == '':
                if resource['download_url'] is not None and resource['download_url'] != '':
                    resource['url'] = resource['download_url']
            if 'format' not in resource \
            or resource['format'] is None \
            or resource['format'] == '':
                fmt = ''
                url_split = resource['url'].split('.')
                if len(url_split) > 1:
                    fmt = url_split[-1]
                resource['format'] = fmt
            resource['format'] = resource['format'].upper()
            if resource['filename'] is None or resource['filename'] == '':
                if resource['url'] is not None and resource['url'] != '':
                    resource['filename'] = resource['url'].split('/')[-1]
                if resource['filename'] is None or resource['filename'] == '':
                    resource['filename'] = resource['name'] + '.' + resource['format'].lower()
            if resource['filename'] is None or resource['filename'] == '':
                resource['filename'] = None
                continue
            if len(resource['filename'].encode('UTF-8')) >= 128:
                resource['filename'] = resource['filename'][:20] \
                                     + '__' + resource['filename'][-20:]
            if resource['filename'].find('/') >= 0:
                resource['filename'] = resource['filename'].replace('/', '_')
            if resource['filename'].find('\n') >= 0:
                resource['filename'] = resource['filename'].replace('\n', '_')
            if len(resource['filename'].split('.')) == 1:
                resource['filename'] += '.' + resource['format'].lower()
            elif resource['filename'].split('.')[-1].upper() != resource['format'].upper():
                resource['filename'] = '.'.join(resource['filename'].split('.')[:-1]) \
                                     + '.' + resource['format'].lower()
        # 復帰
        logger.info('get_all_resources() ended, rc='+str(len(resources)))
        return resources

    def get_valid_resources(self, resources):
        """
        有効なリソース情報を取得する。
        """
        logger = logging.getLogger(__name__)
        logger.info('get_valid_resources() start.')
        # 実行
        if resources is None or not isinstance(resources, list):
            raise Exception('指定されたリソース一覧は異常です。')
        for res in resources:
            if not isinstance(res, dict):
                raise Exception('リソースはdict型で指定して下さい。')
        valid = [v for v in resources if 'name' in v and 'format' in v and 'url' in v]
        for i in range(len(valid)):
            if valid[i] is None or valid[i]['url'] is None:
                valid[i]['format'] = 'XXXXX'
                continue
            if valid[i]['url'].split('.')[-1].lower() == 'html':
                valid[i]['format'] = 'HTML'
        # (1) formatが「CSV、TXT、XLSX、XLS、GeoJSON、HTML」以外は除外する
        valid = [v for v in valid if v['format'] in \
                ['CSV','TXT','XLSX','XLS','GEOJSON','HTML']]
        # (2) １件以下なら終了
        if len(valid) <= 1:
            logger.info('get_valid_resources() ended_1, rc='+str(len(valid)))
            return valid
        # (3) formatがHTMLかつnameが「関連ホームページ」は除外する
        valid = [v for v in valid if not (v['name'] == '関連ホームページ' \
                and v['format'] == 'HTML')]
        # (4) nameに拡張子が設定されている場合は削除する
        for i in range(len(valid)):
            for s in ['.', '_']:
                name_split = valid[i]['name'].split(s)
                if len(name_split) > 1:
                    if name_split[-1].upper() == valid[i]['format'].upper():
                        valid[i]['name'] = valid[i]['name'][:-(len(valid[i]['format'])+1)]
        # (5) nameが同じでformatが異なるものは、優先順「CSV>TXT>XLSX>XLS>GeoJSON>HTML」で１つにする
        uniq_names = set([v['name'] for v in valid])
        if len(valid) > len(uniq_names):
            uniq_resources = []
            uniq_urls = []
            for n in uniq_names:
                f_exists = False
                for f in ['CSV','TXT','XLSX','XLS','GeoJSON','HTML']:
                    for i in range(len(valid)):
                        if valid[i]['name'] != n:
                            continue
                        if valid[i]['format'] != f:
                            continue
                        if valid[i]['url'] is None:
                            continue
                        if valid[i]['url'] in uniq_urls:
                            continue
                        uniq_resources.append(valid[i])
                        uniq_urls.append(valid[i]['url'])
                        f_exists = True
                    if f_exists:
                        break
            valid = uniq_resources
        # 復帰
        logger.info('get_valid_resources() ended, rc='+str(len(valid)))
        return valid

    def get_content_by_resource(self, kind, resource):
        """
        リソースを取得しマップ情報に変換してキャッシュに保存する。
        """
        logger = logging.getLogger(__name__)
        logger.info('get_content_by_resource() start.')
        # 実行
        rc = 0
        content = None
        if kind is None or not isinstance(kind, str) \
        or kind not in [k for k in self.KIND_LIST_NORMALIZED_RE.keys()]:
            msg = 'kindパラメタは無効な値です。'
            print(msg, file=sys.stderr)
            logger.error(msg+'kind='+str(kind))
            raise Exception(msg)
            rc = 1
        if resource is None or not isinstance(resource, dict) \
        or 'url' not in resource or 'filename' not in resource:
            msg = 'resourceパラメタは無効な値です。'
            print(msg, file=sys.stderr)
            logger.error(msg+'resource='+str(resource))
            raise Exception(msg)
            rc = 2
        # リソースの内容を取得する
        if rc == 0:
            content = self.url_get(resource['url'], resource['filename'], \
                    dir=self.package_dir)
        if content is None:
            msg = 'リソース内容の取得に失敗しました。'
            # print(msg, file=sys.stderr)
            logger.error(msg+'url='+str(resource['url']))
            # raise Exception(msg)
            rc = 3
        resource['_content_'] = content

        # 復帰
        logger.info('get_content_by_resource() ended, content='+str(content is not None))
        return resource

    def check_html_page(self, resource, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('check_html_page() start.')
        # 実行
        if resource is None or not isinstance(resource, dict) \
        or '_content_' not in resource:
            msg = 'resourceパラメタは無効な値です。'
            logger.error(msg+str(resource))
            raise Exception(msg)
        if dir is None or not isinstance(dir, str) \
        or dir == '' or not os.path.exists(dir):
            msg = 'dirパラメタは無効な値です。'
            logger.error(msg+str(dir))
            raise Exception(msg)
        if resource['format'].upper() != 'HTML':
            logger.info('check_html_page() ended_1, format!=HTML')
            return resource
        resource['_redirect_'] = None
        cont_str = resource['_content_'].decode('UTF-8', \
                errors='replace').replace('\r\n','\n')
        cont_str_strip = cont_str.strip()
        if len(cont_str_strip) > len(XHTML_PAGE) and cont_str_strip[:len(XHTML_PAGE)] != XHTML_PAGE:
            logger.info('check_html_page() ended_2, content is not HTML')
            return resource
        elif cont_str[:len(LINKDATA_DOWNLOAD_CONTENT)] == LINKDATA_DOWNLOAD_CONTENT:
            resource['_content_'] = cont_str
            resource['_redirect_'] = 'LINKDATA'
            resource = self.check_html_page_linkdata(resource, dir)
        else:
            resource['_content_'] = cont_str
            resource['_redirect_'] = 'DOWNLOAD'
            resource = self.check_html_page_xhtml(resource, dir)
        # 復帰
        logger.info('check_html_page() ended.')
        return resource

    def check_html_page_xhtml(self, resource, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('check_html_page_xhtml() start.')
        # 実行
        redirect = False
        soup = None
        format = None
        re_url = None
        try:
            soup = BeautifulSoup(resource['_content_'], 'html.parser')
            if soup is None:
                logger.warning('chekc_html_page_xhtml() ended_1, content is not HTML.')
                return None
            list_download = soup.select('.list-download')
            if list_download is None or len(list_download) == 0:
                logger.warning('check_html_page_xhtml() ended_2, ' \
                        + 'content not have list-download class.')
                return None
            downloads = list_download[0].select('.download')
            if downloads is None:
                logger.warning('check_html_page_xhtml() ended_3, ' \
                        + 'content not have download class.')
                return None
            for download in downloads:
                if 'data-downloadpath' not in download.attrs \
                or download.attrs['data-downloadpath'] == '':
                    logger.warning('content not have data-downloadpath attribute.')
                    continue
                path = download.attrs['data-downloadpath']
                format = None
                text = download.text.upper()
                for f in ['CSV', 'TEXT', 'TXT', 'TSV', 'XLSX', 'XLS']:
                    if text.find(f) < 0:
                        continue
                    format = f
                    break
                if format is None:
                    continue
                re_url = path
                if re_url[0] == '/':
                    re_url = '/'.join(resource['url'].split('/')[:3]) + path
                break

        except Exception as e:
            logger.exception(e)
            return resource
        finally:
            if soup is not None:
                soup.clear()
                soup = None

        if format is not None and re_url is not None:
            file = re_url.split('/')[-1]
            content = self.url_get(re_url, file, dir)
            if content is None:
                logger.warning('check_html_page_xhtml() ended_4, ' \
                        + 'faild in getting redirect content')
                return None
            resource['url'] = re_url
            resource['filename'] = re_url.split('/')[-1]
            resource['format'] = format
            resource['_content_'] = content
            redirect = True

        # 復帰
        logger.info('check_html_page_xhtml() ended, redirect=' + str(redirect))
        return resource

    def check_html_page_linkdata(self, resource, dir):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('check_html_page_linkdata() start.')
        TYPE_TO_FORMAT = {
                'Table Data(CSV)': 'CSV',
                'Table Data(Text)': 'TEXT',
                'Table Data(Excel)': 'Excel',
                'RDF (Turtle)': None
        }
        # 実行
        redirect = None
        url_base = LINKDATA_URL_BASE
        match = re.search(r"LD.systemUrl = '([^']+)';", resource['_content_'])
        if match is not None:
            url_base = match.group(1)
            if url_base[-1] == '/':
                url_base = url_base[:-1]
        else:
            logger.info('check_html_page_linkdata() ended_2, faild-2')
            # return resource
        attrs = []
        soup = None
        try:
            soup = BeautifulSoup(resource['_content_'], 'html.parser')
            downloads = soup.select('.downloadFileInfoContainer')
            for download in downloads:
                type = download.select('.downloadFileNameAndType')[0].find('div').text.replace('\n','').replace('\t','')
                version = download.select('.downloadFileVersion')[0].find('div').text.replace('\n','').replace('\t','')
                size = download.select('.downloadFileSize')[0].find('div').text.replace('\n','').replace('\t','')
                url = url_base + download.select('.downloadFileDownload')[0].find('a').attrs['href']
                file = url[url.rfind('/')+1:]
                if file.find('/') >= 0:
                    file = file.replace('/', '_')
                if type in TYPE_TO_FORMAT:
                    format = TYPE_TO_FORMAT[type]
                    if format == 'Excel':
                        if file[-4:].upper() == '.XLS':
                            format = 'XLS'
                        elif file[-5:].upper() == '.XLSX':
                            format = 'XLSX'
                        else:
                            # logger.info('check_html_page_linkdata() ended_3, faild-3' \
                            #         + ' ' + str(type) + ' ' + str(file))
                            # return resource
                            format = 'XLSX'
                attrs.append({
                    'format': format,
                    'version': version,
                    'size': size,
                    'url': url,
                    'file': file
                })

        except Exception as e:
            logger.exception(e)
            logger.info('check_html_page_linkdata() ended_4, faild-4')
            return resource
        finally:
            if soup is not None:
                soup.clear()
                soup = None
        logger.debug('attrs=' + str(attrs))

        if len(attrs) == 0:
            logger.info('check_html_page_linkdata() ended_5, faild-5')
            return resource
        for format in ['CSV', 'TEXT', 'TXT', 'TSV', 'XLSX', 'XLS']:
            for attr in attrs:
                if attr['format'] == format:
                    redirect = attr
                    break
            if redirect is not None:
                break
        if redirect is None:
            logger.info('check_html_page_linkdata() ended_6, faild-6')
            return resource
        # リダイレクト先をダウンロードする
        content = self.url_get(redirect['url'], redirect['file'], dir=dir)
        if content is None:
            logger.info('check_html_page_linkdata() ended_7, faild-7')
            return resource
        resource['url'] = redirect['url']
        resource['filename'] = redirect['file']
        resource['format'] = redirect['format']
        resource['_content_'] = content
        # 復帰
        redirect_format = 'None'
        if redirect is not None and 'format' in redirect:
            redirect_format = redirect['format']
        logger.info('check_html_page_linkdata() ended, format=' + redirect_format)
        return resource

    def check_content_format(self, resource, dir):
        """
        コンテンツの内容でformatを検査する
        """
        logger = logging.getLogger(__name__)
        logger.info('check_content_format() start.')
        # 実行
        if resource is None or not isinstance(resource, dict) \
        or 'format' not in resource or not isinstance(resource['format'], str) \
        or 'filename' not in resource or not isinstance(resource['filename'], str):
            msg = 'resourceパラメタは無効な値です。'
            logger.error(msg+str(resource))
            raise Exception(msg)
        if dir is None or not isinstance(dir, str) \
        or dir == '' or not os.path.exists(dir):
            msg = 'dirパラメタは無効な値です。'
            logger.error(msg+str(dir))
            raise Exception(msg)
        format_save = resource['format']
        if type(resource['_content_']) == bytes:
            head_bytes8 = resource['_content_'][:8]
        else:
            head_bytes8 = bytes(str(resource['_content_'])[:8], 'UTF-8')
        if len(resource['_content_']) > 8 and head_bytes8 == CONTENT_HEADER_XLS:
            # XLS形式：DOC/XLS/PPTも同じだが判定省略
            if resource['format'].upper() != 'XLS':
                logger.warning('format変更1, ' + resource['format'] + ' -> XLS')
                resource['format'] = 'XLS' 
        elif len(resource['_content_']) > 6 and head_bytes8[:6] == CONTENT_HEADER_XLSX:
            # XLSX形式：ZIP/DOCX/XLSX/PPTXも同じだが判定省略
            if resource['format'].upper() != 'XLSX': 
                logger.warning('format変更2, ' + resource['format'] + ' -> XLSX')
                resource['format'] = 'XLSX'
        else:
            if type(resource['_content_']) == bytes:
                if b'{' in resource['_content_'][0:4] \
                and  b'}' in resource['_content_'][-3:]:
                    # JSON形式(そのまま) 
                    if resource['format'] != 'GeoJSON':
                        logger.warning('format変更3, ' + resource['format'] + ' -> GeoJSON')
                        resource['format'] = 'GeoJSON'
                elif resource['_content_'].count(b'\n') \
                        <= resource['_content_'].count(b','):
                    # CSV形式 
                    if resource['format'].upper() != 'CSV':
                        logger.warning('format変更4, ' + resource['format'] + ' -> CSV')
                        resource['format'] = 'CSV' 
                elif resource['_content_'].count(b'\n') \
                        <= resource['_content_'].count(b'\t'):
                    # TSV形式：TXT/TEXTも同じ
                    if resource['format'].upper() not in ['TXT', 'TSV', 'TEXT']:
                        logger.warning('format変更5, ' + resource['format'] + ' -> TXT')
                        resource['format'] = 'TXT'
                else:
                    # 指定のまま
                    logger.warning('format不明, ' + resource['format'])
            else:
                if '{' in resource['_content_'][0:4] \
                and  '}' in resource['_content_'][-3:]:
                    # JSON形式(そのまま)
                    if resource['format'] != 'GeoJSON':
                        logger.warning('format変更6, ' + resource['format'] + ' -> GeoJSON')
                        resource['format'] = 'GeoJSON'
                elif resource['_content_'].count('\n') \
                        <= resource['_content_'].count(','):
                    # CSV形式 
                    if resource['format'].upper() != 'CSV':
                        logger.warning('format変更7, ' + resource['format'] + ' -> CSV')
                        resource['format'] = 'CSV' 
                elif resource['_content_'].count('\n') \
                        <= resource['_content_'].count('\t'):
                    # TSV形式：TXT/TEXTも同じ
                    if resource['format'].upper() not in ['TXT', 'TSV', 'TEXT']:
                        logger.warning('format変更8, ' + resource['format'] + ' -> TXT')
                        resource['format'] = 'TXT'
                else:
                    # 指定のまま
                    logger.warning('format不明, ' + resource['format'])

        # Excel形式でファイルの拡張子が違っている場合は変更する
        if resource['format'] in ['XLSX', 'XLS'] \
        and resource['filename'].split('.')[-1].upper() != resource['format'].upper():
            dist_name = '.'.join(resource['filename'].split('.')[:-1]) \
                    + '.' + resource['format'].lower()
            os.rename(os.path.join(dir, resource['filename']),
                    os.path.join(dir, dist_name))
            resource['filename'] = dist_name
            dataset_path = os.path.join(dir, resource['filename'])

        # 必要ならコード変換を行う
        if resource['format'] in ['CSV', 'TEXT', 'TSV', 'TXT', 'GeoJSON']: 
            content = self.convert_to_utf8(resource['_content_'])
            if content is None:
                logger.error('content format error, format=' + resource['format'])
            resource['_content_'] = content

        # 復帰
        logger.info('check_content_format() ended, '+format_save+'->'+resource['format'])
        return resource

    def content_to_table(self, resource, dir):
        """
        コンテンツを表形式に変換する
        """
        logger = logging.getLogger(__name__)
        logger.info('content_to_table() start.')
        # 実行
        if resource is None or not isinstance(resource, dict) \
        or 'format' not in resource or not isinstance(resource['format'], str) \
        or 'filename' not in resource or not isinstance(resource['filename'], str) \
        or '_content_' not in resource or resource['_content_'] is None:
            msg = 'resourceパラメタは無効な値です。'
            logger.error(msg+str(resource))
            raise Exception(msg)
        if dir is None or not isinstance(dir, str) \
        or dir == '' or not os.path.exists(dir):
            msg = 'dirパラメタは無効な値です。'
            logger.error(msg+str(dir))
            raise Exception(msg)
        resource['_table_'] = None
        table = None
        if resource['format'] == 'CSV':
            # CSV形式
            logger.debug('format is ' + resource['format'])
            table = self.table_from_csv(resource['_content_'])
        elif resource['format'] in ['TEXT', 'TXT', 'TSV']:
            # テキスト形式(TSV)
            logger.debug('format is ' + resource['format'])
            table = self.table_from_tsv(resource['_content_'])
        elif resource['format'] == 'XLS':
            # Excel形式(XLS)
            logger.debug('format is ' + resource['format'])
            table = self.table_from_xls(os.path.join(dir, resource['filename']))
        elif resource['format'] == 'XLSX':
            # Excel形式(XLSX) 
            logger.debug('format is ' + resource['format'])
            table = self.table_from_xlsx(os.path.join(dir, resource['filename']))
        elif resource['format'] == 'GeoJSON':
            # GeoJSON形式(JSON)
            logger.debug('format is ' + resource['format'])
            table = self.table_from_GeoJSON(resource['_content_'])
        else:
            # 未サポートの形式
            logger.error('ERROR in format, format=' + resource['format'])
            # raise Exception('ERROR in format, format=' + resource['format'])
        resource['_table_'] = table
        # 復帰
        logger.info('content_to_table() ended.')
        return resource

    def make_map_from_table(self, resource, kind):
        """
        表形式の項目名からマップ情報を作成する
        """
        logger = logging.getLogger(__name__)
        logger.info('make_map_from_table() start.')
        # 実行
        if resource is None or not isinstance(resource, dict) \
        or '_table_' not in resource or not isinstance(resource['_table_'], list):
            msg = 'resourceパラメタは無効な値です。' 
            logger.error(msg+str(resource))
            raise Exception(msg)
        if kind is None or not isinstance(kind, str) \
        or kind == '' or kind not in self.KIND_LIST_NORMALIZED_RE:
            msg = 'kindパラメタは無効な値です。'
            logger.error(msg+str(kind))
            raise Exception(msg)

        table = resource['_table_']
        map_info = {
            "kind": resource['_package_']['_info_']['kind'],
            "dataset": resource['_package_']['_info_']['title'],
            "lat": -1,
            "lng": -1,
            "name": [-1],
            "id": -1,
            "info": [-1],
            "address": -1,
            "header": -1
        }
        map_msg = ''
        title_rows = [list(row) for row in table[:min(HEADER_ROWS , len(table))]]
        for i_row in range(len(title_rows)):
            for i_col in range(len(title_rows[i_row])):
                if title_rows[i_row][i_col] is None:
                    title_rows[i_row][i_col] = ''
        for t in range(len(title_rows)):
            if title_rows[t] == []:
                continue
            address = [i for i in range(len(title_rows[t])) \
                    if title_rows[t][i] in self.address_item_list]
            if len(address) == 0:
                # 住所項目なし
                address = [-1]
            address = address[0]
            lat = [i for i in range(len(title_rows[t])) \
                    if title_rows[t][i] in self.lat_item_list \
                    or (type(title_rows[t][i])==str \
                      and len(title_rows[t][i])>4 \
                      and title_rows[t][i][-4:]=='#lat')]
            if len(lat) == 0:
                # 緯度項目なし
                lat = [-1]
            lat = lat[0]
            lng = [i for i in range(len(title_rows[t])) \
                    if title_rows[t][i] in self.lng_item_list \
                    or (type(title_rows[t][i])==str \
                      and len(title_rows[t][i])>5 \
                      and title_rows[t][i][-5:]=='#long')]
            if len(lng) == 0:
                # 経度項目なし
                lng = [-1]
            lng = lng[0]
            name = [i for i in range(len(title_rows[t])) \
                    if title_rows[t][i] in self.name_item_list_1 \
                    or (type(title_rows[t][i])==str \
                      and len(title_rows[t][i])>6 \
                      and title_rows[t][i][-6:]=='#label')]
            if len(name) > 0:
                name = [name[0]]
            else:
                if kind in self.name_item_list_2:
                    name = [title_rows[t].index(w) for w in self.name_item_list_2[kind] \
                            if w in title_rows[t]]
                if len(name) == 0:
                    # 名称項目なし
                    name = [-1]
                    if '種別' in title_rows[t] and '住所' in title_rows[t]:
                        name = [title_rows[t].index('種別'), title_rows[t].index('住所')]
            id = [i for i in range(len(title_rows[t])) \
                    if title_rows[t][i] in self.id_item_list]
            if len(id) == 0:
                # ID項目なし
                id = [-1]
            id = id[0]

            map_info['lat'] = lat
            map_info['lng'] = lng
            map_info['name'] = name
            map_info['id'] = id
            map_info['address'] = address
            map_info['header'] = t + 1
            used_index = [map_info['lat'],map_info['lng']]	# ,map_info['id']]
            # used_index.extend(map_info['name'])
            map_info['info'] = [i for i in range(len(title_rows[map_info['header']-1])) \
                    if i not in used_index]
            logger.debug('map_info=' + str(map_info))
            if ((map_info['lat'] >= 0 and map_info['lng'] >= 0) \
              or map_info['address'] >= 0) \
            and map_info['name'] != [-1]:
                # 必須項目が揃っている
                break
            # 必須項目なし
            # raise Exception('必須項目なし')
        # 最終確認
        if (map_info['lat'] < 0 or map_info['lng'] < 0) \
        and map_info['address'] < 0 \
        or map_info['name'] == [-1]:
            # 必須項目なし
            map_msg = str({k:v for k,v in map_info.items() \
                           if k in ['id','name','lat','lng','address']})
            logger.info('必須項目なし, '+str(map_info))
            map_info = {}
        else:
            logger.debug('headers=' + str(title_rows[map_info['header']-1]))
            # ToDo:削除 info.append(map_info)
        resource['_map_'] = map_info
        resource['_map_msg_'] = map_msg

        # 復帰
        logger.info('make_map_from_table() ended, map=' + str(map_info))
        return resource

    def get_package_kind(self, package):
        """
        パッケージの正規化種別を返却する
        """
        return package['_info_']['kind']

    def data_from_content(self, content, info, file):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('data_from_content() start.')
        data = None
        if info['format'] == 'CSV':
            data = self.data_from_csv(content, info);
        elif info['format'] in ['TEXT', 'TXT', 'TSV']:
            data = self.data_from_tsv(content, info);
        elif info['format'] == 'XLS':
            data = self.data_from_xls(content, info, file);
        elif info['format'] == 'XLSX':
            data = self.data_from_xlsx(content, info, file);
        # 復帰
        logger.info('data_from_content() ended.')
        return data

    def table_from_csv(self, content):
        """
        CSV形式のデータをテーブル形式に変換する。
        :param content: str型、CSV形式文字列
        :return: list型、list形式行のlist形式
        """
        logger = logging.getLogger(__name__)
        logger.info('table_from_csv() start.')
        # 実行
        table = [v for v in csv.reader(content.replace('\r\n','\n').replace('\r','\n').split('\n'))]
        # 復帰
        logger.info('table_from_csv() ended, rc=' + str(len(table)) \
                + ' table[:5]=' + str(table[:5]))
        return table

    def table_from_tsv(self, content):
        """
        """
        logger = logging.getLogger(__name__)
        logger.info('table_from_tsv() start.')
        # 実行
        table = [v for v in csv.reader( \
                content.replace('\r\n', '\n').replace('\r','\n').split('\n'),
                delimiter='\t', quotechar=None)]
        if len(table) > 0 and len(table[0]) > 0 \
        and type(table[0][0]) == str and table[0][0] == '#LINK':
            # LinkTable形式
            logger.debug('#LINK形式')
            property = True
            link_data = []
            for i in range(len(table)):
                if len(table[i]) == 0 or len(table[i][0]) == 0:
                    continue;
                if property:
                    if table[i][0][0] == '#':
                        if table[i][0] == '#property':
                            properties = [v[0] for v in table[i+1:] \
                                    if len(v)>0 and v[0][0]!='#' \
                                    and not re.match('^[0-9a-zA-Z\\:\\+\\-\\._#]{0,18}$', v[0])]
                            logger.debug('properties=' + str(properties))
                            if len(properties) == 0:
                                link_data.append(['id'] + table[i][1:])
                            else:
                                link_data.append(['名称'] + table[i][1:])
                        continue
                    else:
                        property = False
                link_data.append(table[i])
            table = link_data
        # 復帰
        logger.info('table_from_tsv() ended, rc=' + str(len(table)) \
                + ', table[:5]=' + str(table[:5]))
        return table

    def table_from_xls(self, file, rows=0):
        """
        Excel形式(XLS)をテーブル形式に変換する。
        """
        logger = logging.getLogger(__name__)
        logger.info('table_from_xls() start, rows=' + str(rows))
        # 実行
        table = []
        try:
            excel_book = xlrd.open_workbook(file)
            excel_sheet = excel_book.sheet_by_index(0)
            for row in range(excel_sheet.nrows):
                row_value = []
                for col in range(excel_sheet.ncols):
                    v = excel_sheet.cell(row, col).value
                    if v is None:
                        v = ''
                    elif isinstance(v, datetime.time):
                        v = str(v)
                        if v[-3:] == ':00':
                            v = v[:-3]
                    elif isinstance(v, datetime.datetime):
                        v = str(v)
                        if v[-9:] == ' 00:00:00':
                            v = v[:-9]
                        elif v[-3:] == ':00':
                            v = v[:-3]
                    row_value.append(v)
                table.append(row_value)
                if rows > 0  and len(table) >= rows:
                    break
            excel_book = None
            # logger.debug('table=' + str(table).replace('], [', '],\n['))
            logger.debug('table[:3]=' + str(table[:3]))
        except Exception as e:
            logger.error('error in Excel(XLS) data, file=' + file)
            logger.exception(e)
            table = []
            print('Excel形式(XLS)誤り：' + file, file=sys.stderr, end='')
        # 復帰
        logger.info('table_from_xls() ended, rc=' + str(len(table)) \
                + ', table[:5]=' + str(table[:5]))
        return table

    def table_from_xlsx(self, file, rows=0):
        """
        Excel形式(XLSX)をテーブル形式に変換する。
        """
        logger = logging.getLogger(__name__)
        logger.info('table_from_xlsx() start, rows=' + str(rows))
        # 実行
        table = []
        try:
            excel_book = openpyxl.load_workbook(file, \
                            keep_vba=False, read_only=True, data_only=True)
            excel_sheet = excel_book.worksheets[0]
            for row in excel_sheet.iter_rows():
                row_value = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        v = ''
                    elif isinstance(v, datetime.time):
                        v = str(v)
                        if v[-3:] == ':00':
                            v = v[:-3]
                    elif isinstance(v, datetime.datetime):
                        v = str(v)
                        if v[-9:] == ' 00:00:00':
                            v = v[:-9]
                        elif v[-3:] == ':00':
                            v = v[:-3]
                    elif isinstance(v, datetime.timedelta):
                        v = str(v)
                        if v == '1 day':
                            v = '24:00'
                    row_value.append(v)
                table.append(row_value)
                if rows > 0  and len(table) >= rows:
                    break
            excel_book.close()
            # logger.debug('table=' + str(table).replace('], [', '],\n['))
            logger.debug('table[:3]=' + str(table[:3]))
        except Exception as e:
            logger.error('error in Excel data(XLSX), file=' + file)
            logger.exception(e)
            table = []
            print('Excel形式(XLSX)誤り：' + file, file=sys.stderr, end='')
        # 復帰
        logger.info('table_from_xlsx() ended, rc=' + str(len(table)) \
                + ', table[:5]=' + str(table[:5]))
        return table

    def table_from_GeoJSON(self, content):
        """
        GeoJSON形式(JSON)をテーブル形式に変換する。
        """
        logger = logging.getLogger(__name__)
        logger.info('table_from_GeoJSON() start.')
        # 実行
        table = []
        gj = json.loads(content)
        if 'type' not in gj or gj['type'] != 'FeatureCollection' \
        or 'features' not in gj or not isinstance(gj['features'], list):
            logger.info('table_from_GeoJSON() ended_1, table=[]')
            return table
        # 項目名を抽出
        items = []
        for feature in gj['features']:
            if 'type' not in feature or feature['type'] != 'Feature' \
            or 'geometry' not in feature or not isinstance(feature['geometry'], dict) \
            or 'type' not in feature['geometry'] or feature['geometry']['type'] != 'Point' \
            or 'coordinates' not in feature['geometry'] \
            or not isinstance(feature['geometry']['coordinates'], list) \
            or len(feature['geometry']['coordinates']) < 2:
                continue
            for k in feature['properties'].keys():
                if k not in items:
                    items.append(k)
        if len(items) == 0:
            logger.info('table_from_GeoJSON() ended_2, table=[]')
            return table
        items.extend(['緯度', '経度'])
        table.append(items)
        # データを抽出
        for feature in gj['features']:
            if 'type' not in feature or feature['type'] != 'Feature' \
            or 'geometry' not in feature or not isinstance(feature['geometry'], dict) \
            or 'type' not in feature['geometry'] or feature['geometry']['type'] != 'Point' \
            or 'coordinates' not in feature['geometry'] \
            or not isinstance(feature['geometry']['coordinates'], list) \
            or len(feature['geometry']['coordinates']) < 2:
                continue
            row = [''] * (len(items) - 2)
            for i, item in enumerate(items[:-3]):
                if item in feature['properties']:
                    row[i] = feature['properties'][item]
            row.extend([feature['geometry']['coordinates'][0], \
                    feature['geometry']['coordinates'][1]])
            table.append(row)

        # 復帰
        logger.info('table_from_GeoJSON() ended, rc=' + str(len(table)) \
                + ', table[:5]=' + str(table[:5]))
        return table

def setup_logger(name, level, log_file='crawler.log', log_dir='log'):
    """
    コマンド実行時のログ初期化
    :param name: str型、関数、__main__
    :param level: int型、logging.INFO、logging.DEBUG、...
    :param log_file: str型/stderr、ログファイル名またはstderr
    :param log_dir: str型、ログディレクトリ名、log_fileがログファイル名の場合に有効なパラメタ
    :return: logger
    """
    logger = logging.getLogger(name)
    logger.parent.setLevel(level)
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    log_handler_format = logging.Formatter(log_format)
    if type(log_file) == str:
        # ファイル出力ハンドラ
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        log_file_handler = logging.FileHandler(os.path.join(log_dir, log_file))
        # log_file_handler= logging.handlers.RotatingFileHandler(os.path.join(log_dir, log_file), maxBytes=1024000, backupCount=5)
        log_file_handler.setLevel(level)
        log_file_handler.setFormatter(log_handler_format)
        logger.addHandler(log_file_handler)
    else:
        # 標準エラー出力ハンドラ
        log_stream_handler = logging.StreamHandler()
        log_stream_handler.setLevel(level)
        log_stream_handler.setFormatter(log_handler_format)
        logger.addHandler(log_stream_handler)
    return logger

# コマンド呼び出し
if __name__ == '__main__':
    # 初期化
    rc = 0
    cobj = None
    logger = setup_logger(__name__, logging.DEBUG)  # logging.INFO
    # 実行
    try:
        # パラメタチェック
        import argparse
        p = argparse.ArgumentParser()
        p.add_argument('-l', '--site_list', action='store_true',
                help='定義されているサイト（自治体名）を出力する。')
        p.add_argument('names', nargs='*', type=str,
                help='自治体名[/パッケージID[/...]]を指定する。省略した場合は全自治体を対象とする。')
        args = p.parse_args(sys.argv[1:])
        names = args.names
        # 開始
        msg = 'crawler.py start.'
        logger.info(msg)
        print(msg, file=sys.stderr)
        # 取得開始
        cobj = Crawler()
        names_all = cobj.get_names()
        if args.site_list:
            print('定義済自治体名：' + str(names_all), file=sys.stderr)
        else:
            if len(names) == 0:
                # 全サイトを処理する
                names = names_all
            rc = cobj.crawler_main(names)

    except Exception as e:
        logger.exception(e)
        print('', file=sys.stderr)
        rc = 99

    finally:
        cobj = None

    # 終了
    msg = 'crawler.py ended, rc=' + str(rc)
    logger.info(msg)
    print(msg, file=sys.stderr)

    # 復帰
    sys.exit(rc)

